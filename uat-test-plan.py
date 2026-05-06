#!/usr/bin/env python3
"""Generate Yannis EOSE UAT Test Plan Excel file."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime

wb = openpyxl.Workbook()

# ── Styles ──────────────────────────────────────────────────────────────
HEADER_FILL = PatternFill("solid", fgColor="1F2937")
HEADER_FONT = Font(name="Inter", bold=True, color="FFFFFF", size=11)
SECTION_FILL = PatternFill("solid", fgColor="374151")
SECTION_FONT = Font(name="Inter", bold=True, color="FFFFFF", size=11)
SUBSECTION_FILL = PatternFill("solid", fgColor="6B7280")
SUBSECTION_FONT = Font(name="Inter", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Inter", size=10)
BOLD_FONT = Font(name="Inter", size=10, bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

# Status dropdown colors
PASS_FILL = PatternFill("solid", fgColor="D1FAE5")
FAIL_FILL = PatternFill("solid", fgColor="FEE2E2")
PENDING_FILL = PatternFill("solid", fgColor="FEF3C7")
BLOCKED_FILL = PatternFill("solid", fgColor="E0E7FF")
NOT_TESTED_FILL = PatternFill("solid", fgColor="F3F4F6")

COLUMNS = [
    ("A", 6,  "ID"),
    ("B", 14, "Module"),
    ("C", 20, "Sub-Module"),
    ("D", 50, "Test Case"),
    ("E", 35, "Expected Result"),
    ("F", 14, "Priority"),
    ("G", 14, "Tested By Role"),
    ("H", 12, "Status"),
    ("I", 30, "Actual Result / Notes"),
    ("J", 14, "Tested By"),
    ("K", 12, "Date Tested"),
    ("L", 14, "Defect ID"),
]


def setup_sheet(ws, title):
    ws.title = title
    ws.sheet_properties.tabColor = "4F46E5"
    # Column widths
    for col_letter, width, _ in COLUMNS:
        ws.column_dimensions[col_letter].width = width
    # Header row
    for i, (_, _, label) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=i, value=label)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:L1"
    # Status validation
    dv = DataValidation(
        type="list",
        formula1='"PASS,FAIL,PENDING,BLOCKED,NOT TESTED"',
        allow_blank=True,
    )
    dv.error = "Pick a status"
    dv.errorTitle = "Invalid Status"
    ws.add_data_validation(dv)
    return dv


def add_section(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER
    for c in range(2, len(COLUMNS) + 1):
        ws.cell(row=row, column=c).fill = SECTION_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    return row + 1


def add_subsection(ws, row, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = SUBSECTION_FONT
    cell.fill = SUBSECTION_FILL
    cell.alignment = Alignment(vertical="center")
    cell.border = THIN_BORDER
    for c in range(2, len(COLUMNS) + 1):
        ws.cell(row=row, column=c).fill = SUBSECTION_FILL
        ws.cell(row=row, column=c).border = THIN_BORDER
    return row + 1


def add_row(ws, row, dv, tc_id, module, sub, case, expected, priority, role):
    vals = [tc_id, module, sub, case, expected, priority, role, "NOT TESTED", "", "", "", ""]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.font = BODY_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER
    # Apply validation to status column
    status_cell = f"H{row}"
    dv.add(status_cell)
    ws.cell(row=row, column=8).fill = NOT_TESTED_FILL
    return row + 1


# ════════════════════════════════════════════════════════════════════════
# TEST CASES — ordered by the priority flow the user requested
# ════════════════════════════════════════════════════════════════════════

test_cases = []
tc = 0

def T(module, sub, case, expected, priority, role):
    global tc
    tc += 1
    test_cases.append((f"TC-{tc:04d}", module, sub, case, expected, priority, role))

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 1: AUTH & SETUP
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_AUTH = "1. AUTHENTICATION & INITIAL SETUP"

T("Auth", "Initial Setup", "Navigate to /auth/setup when no users exist — create SuperAdmin", "SuperAdmin account created, redirected to /admin", "P0 - Critical", "SuperAdmin")
T("Auth", "Login", "Login with valid email/password", "Redirected to role-appropriate dashboard", "P0 - Critical", "All Roles")
T("Auth", "Login", "Login with invalid credentials", "Error message shown, no redirect", "P0 - Critical", "All Roles")
T("Auth", "Login", "Login with deactivated account", "Access denied message shown", "P1 - High", "All Roles")
T("Auth", "Forgot Password", "Submit forgot-password with valid email", "Reset email sent, confirmation shown", "P1 - High", "All Roles")
T("Auth", "Reset Password", "Follow reset link and set new password", "Password updated, can login with new password", "P1 - High", "All Roles")
T("Auth", "Logout", "Click logout from any page", "Session cleared, redirected to login", "P1 - High", "All Roles")
T("Auth", "Session", "Session expires after configured timeout", "User redirected to login", "P2 - Medium", "All Roles")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 2: PRODUCTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_PRODUCTS = "2. PRODUCTS MANAGEMENT"

T("Products", "Create", "Create product with all required fields (name, SKU, cost price, selling price, category)", "Product created with UUIDv7 PK, appears in list", "P0 - Critical", "Stock Manager")
T("Products", "Create", "Create product with optional initial stock (quantity + location)", "Product created AND FIFO batch created at specified location", "P0 - Critical", "Stock Manager")
T("Products", "Create", "Create product with gallery images (upload to S3/R2)", "Images uploaded, URLs stored, displayed in product detail", "P1 - High", "Stock Manager")
T("Products", "Create", "Attempt product creation without required fields", "Validation errors shown per field", "P1 - High", "Stock Manager")
T("Products", "List", "View products list page with pagination", "Products listed with correct pagination, search works", "P0 - Critical", "Stock Manager")
T("Products", "List", "Search products by name or SKU", "Filtered results returned", "P1 - High", "Stock Manager")
T("Products", "Detail", "View product detail page — cost fields visible to finance roles", "Cost/margin fields shown for SuperAdmin/Finance, stripped for others", "P0 - Critical", "SuperAdmin / CS Agent")
T("Products", "Edit", "Update product name, price, description", "Changes saved, audit trail shows old/new values", "P1 - High", "Stock Manager")
T("Products", "Edit", "Update product cost price", "Cost updated, FIFO batches unaffected (only new batches use new cost)", "P1 - High", "Stock Manager")
T("Products", "Archive", "Stock Manager requests product archive", "Permission request created (PENDING), product still active", "P1 - High", "Stock Manager")
T("Products", "Archive", "SuperAdmin approves archive request", "Product archived, no longer appears in active product list", "P1 - High", "SuperAdmin")
T("Products", "CLS", "Non-finance user views product — cost_price, landed_cost, margin hidden", "Fields stripped from API response, not in DOM", "P0 - Critical", "CS Agent")
T("Products", "Categories", "Create product category", "Category created, available in product create form", "P2 - Medium", "Stock Manager")
T("Products", "Categories", "Edit product category", "Category updated", "P2 - Medium", "Stock Manager")
T("Products", "Offer Templates", "Create offer template for a product (price tiers)", "Template created, available for campaign forms", "P1 - High", "HoM / Stock Mgr")
T("Products", "Offer Templates", "Edit offer template (change tier pricing)", "Template updated, edge forms reflect new pricing", "P1 - High", "HoM / Stock Mgr")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 3: WAREHOUSES / LOCATIONS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_WAREHOUSES = "3. WAREHOUSES & LOGISTICS LOCATIONS"

T("Warehouses", "Create", "Create a Yannis-owned warehouse location", "Location created, appears in inventory location list", "P0 - Critical", "Stock Manager")
T("Warehouses", "List", "View all warehouse locations", "All locations listed with stock counts", "P0 - Critical", "Stock Manager")
T("Logistics Locations", "Create Provider", "Create a 3PL logistics provider company", "Provider created with name, contact info", "P0 - Critical", "HoLogistics")
T("Logistics Locations", "Create Location", "Create a 3PL location under a provider (with optional WhatsApp group link)", "Location created, linked to provider", "P0 - Critical", "HoLogistics")
T("Logistics Locations", "Edit", "Update location details (name, WhatsApp link)", "Changes saved, audit trail updated", "P1 - High", "HoLogistics")
T("Logistics Locations", "WhatsApp Link", "Add WhatsApp group link — only chat.whatsapp.com or wa.me URLs accepted", "Valid URLs saved; invalid rejected", "P1 - High", "HoLogistics")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 4: SHIPMENTS (INBOUND)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_SHIPMENTS = "4. INBOUND SHIPMENTS"

T("Shipments", "Create", "Create shipment with multiple product lines (expected quantities, factory costs)", "Shipment created in CREATED status, reference SHIP-YYYY-XXXX generated", "P0 - Critical", "Stock Manager")
T("Shipments", "Create Arrived", "Create shipment with 'arrived now' flag", "Shipment created directly in ARRIVED status", "P1 - High", "Stock Manager")
T("Shipments", "List", "View shipments list with status tabs", "Shipments listed, filterable by status", "P0 - Critical", "Stock Manager")
T("Shipments", "Detail", "View shipment detail page with line items", "All lines shown with expected vs received quantities", "P0 - Critical", "Stock Manager")
T("Shipments", "Mark In Transit", "Transition CREATED -> IN_TRANSIT", "Status updated, visible in pipeline view", "P1 - High", "Stock Manager")
T("Shipments", "Mark Arrived", "Transition CREATED/IN_TRANSIT -> ARRIVED", "arrived_at stamped, lines unlock for received qty entry", "P0 - Critical", "Stock Manager")
T("Shipments", "Enter Received Qty", "Enter received quantities per line (may differ from expected)", "Quantities saved; variance_reason required when received != expected", "P0 - Critical", "Stock Manager")
T("Shipments", "Verify", "Verify ARRIVED shipment — creates stock batches + inventory levels", "FIFO batches created, inventory_levels.stock_count incremented, stock_movements logged (INTAKE), landing cost allocated across lines", "P0 - Critical", "Stock Manager")
T("Shipments", "Verify - Cost Allocation", "Verify shipment with total_landing_cost — cost allocated by received_qty x factory_cost weight", "Each batch has correct landed cost proportional to line value", "P1 - High", "Stock Manager")
T("Shipments", "Verify - Variance", "Verify with received < expected without variance_reason", "Server rejects — variance_reason required", "P1 - High", "Stock Manager")
T("Shipments", "Close", "Close a VERIFIED shipment", "Status becomes CLOSED, row immutable", "P1 - High", "Stock Manager")
T("Shipments", "Cancel", "Cancel a CREATED/IN_TRANSIT/ARRIVED shipment with reason (>=10 chars)", "Status CANCELLED, no inventory side effects", "P1 - High", "Stock Manager")
T("Shipments", "Cancel Verified", "Attempt to cancel a VERIFIED shipment", "Server rejects — verified shipments cannot be cancelled", "P1 - High", "Stock Manager")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 5: STOCK & INVENTORY FLOWS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_STOCK = "5. STOCK & INVENTORY FLOWS"

T("Inventory", "Stock Intake", "Single-product stock intake at a warehouse (qty + cost)", "FIFO batch created, inventory_levels updated, INTAKE movement logged", "P0 - Critical", "Stock Manager")
T("Inventory", "View Levels", "View inventory levels page — per product x location", "All locations and products shown with stock_count, reserved_count", "P0 - Critical", "Stock Manager")
T("Inventory", "Level Detail", "View inventory detail — batches + movements for a product at a location", "FIFO batches listed (remaining_qty, cost), movements chronologically", "P1 - High", "Stock Manager")
T("Inventory", "Transfer Initiate", "Initiate stock transfer from Main Warehouse to 3PL location", "Transfer created, stock IN_TRANSIT, source inventory decremented", "P0 - Critical", "Stock Manager")
T("Inventory", "Transfer Verify", "3PL Manager / HoLogistics verifies transfer receipt (actual qty)", "Destination inventory incremented, TRANSFER movement logged", "P0 - Critical", "TPL Manager / HoLogistics")
T("Inventory", "Transfer Shrinkage", "Verify transfer with received < sent qty", "Shrinkage alert generated to CEO + HoLogistics", "P1 - High", "TPL Manager")
T("Inventory", "Adjustment Increase", "Stock adjustment — increase quantity at a location", "inventory_levels updated, ADJUSTMENT movement logged with reason", "P1 - High", "Stock Manager")
T("Inventory", "Adjustment Decrease", "Stock adjustment — decrease quantity at a location", "inventory_levels updated, ADJUSTMENT movement logged with reason", "P1 - High", "Stock Manager")
T("Inventory", "Low Stock Alerts", "View low stock alerts when any product falls below threshold", "Alert shown for products below reorder level", "P1 - High", "Stock Manager")
T("Inventory", "Reconciliation Create", "Create stock reconciliation form (physical count vs digital, mandatory reason code)", "Reconciliation created with reason: Damaged/Lost/Expired/Theft", "P1 - High", "Stock Manager / TPL Mgr")
T("Inventory", "Reconciliation Resolve", "Resolve reconciliation — adjusts stock to match physical count", "Inventory adjusted, movements logged, audit trail captured", "P1 - High", "Stock Manager")
T("Inventory", "Ghost Stock Lock", "Physical count mismatch locks dispatch at that location", "Dispatch button disabled until reconciliation submitted", "P0 - Critical", "Stock Manager")
T("Inventory", "Virtual Buffer", "Sales form sees 10% less stock than actual (virtual buffer)", "Edge form shows 'Sold Out' when available < 10% buffer", "P1 - High", "System (Edge)")
T("Inventory", "FIFO Ordering", "When multiple batches exist, oldest batch consumed first on delivery", "Batch A ($5/unit) consumed before Batch B ($7/unit)", "P0 - Critical", "System")
T("Inventory", "Returned Stock Restock", "Return marked as sellable at 3PL → local restock (no freight back)", "3PL local inventory incremented, RESTOCK movement logged", "P1 - High", "TPL Manager")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 6: MARKETING / ORDER ENTRY
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_MARKETING = "6. MARKETING — CAMPAIGNS, FORMS, FUNDING, AD SPEND"

# Campaigns & Forms
T("Marketing", "Create Campaign", "Create campaign/form with product, offer template, form config", "Campaign created, deployable via 3 modes (Shadow DOM, iframe, hosted)", "P0 - Critical", "Media Buyer / HoM")
T("Marketing", "Edit Campaign", "Edit campaign details (heading, subtitle, button text, accent color)", "Changes saved, preview updated", "P1 - High", "Media Buyer / HoM")
T("Marketing", "Form Deploy - Shadow DOM", "Deploy form as Shadow DOM snippet — copy embed code", "Embed code copied, form renders on external page", "P1 - High", "Media Buyer")
T("Marketing", "Form Deploy - Iframe", "Deploy form as iframe — copy iframe embed", "Iframe embed works on external page", "P1 - High", "Media Buyer")
T("Marketing", "Form Deploy - Hosted", "Deploy form as hosted URL", "Hosted URL accessible, form functional", "P1 - High", "Media Buyer")
T("Marketing", "Form Config", "Configure form with successCallbackUrl (optional)", "When set: redirects to thank-you page after submit; when empty: shows inline success", "P1 - High", "Media Buyer")

# Edge Form Submission
T("Marketing", "Edge Form Submit", "Submit order via edge form (Cloudflare Worker)", "Order created as UNPROCESSED, response < 400ms", "P0 - Critical", "Public (Customer)")
T("Marketing", "Edge Dedup Same MB", "Same MB submits same phone+product within 6 hours", "KV short-circuits — alreadySubmitted:true, submit button disabled, no redirect", "P0 - Critical", "System (Edge)")
T("Marketing", "Edge Dedup Cross MB", "Different MB submits same phone+product within 6 hours", "API returns crossFunnelAttempt:true, cross_funnel_attempts row inserted, no order created", "P0 - Critical", "System (Edge)")
T("Marketing", "Edge Circuit Breaker", "API latency > 2000ms or 5xx — order buffered in QStash", "Order buffered, user sees 'Order received, processing shortly'", "P1 - High", "System (Edge)")
T("Marketing", "Edge Sold Out", "Submit when stock < 10% buffer (virtual buffer)", "Form shows 'Sold Out' response", "P1 - High", "System (Edge)")
T("Marketing", "Paystack Payment", "Submit order via form with Paystack payment flow", "Redirect to Paystack, on success redirect to /payment/thank-you, order created", "P1 - High", "Public (Customer)")

# Funding
T("Marketing", "Send Funding", "HoM sends funding to Media Buyer (amount + receipt image)", "Funding record created (SENT), MB gets push notification", "P0 - Critical", "HoM")
T("Marketing", "Mark Received", "Media Buyer marks funding as received", "Status → COMPLETED", "P0 - Critical", "Media Buyer")
T("Marketing", "Dispute Funding", "Media Buyer marks 'Not Received'", "Status → DISPUTED, alert to CEO + HoM", "P1 - High", "Media Buyer")
T("Marketing", "Request Funding (MB)", "Media Buyer requests funding from HoM", "Request created (PENDING), HoM notified (NOT SuperAdmin/Finance)", "P0 - Critical", "Media Buyer")
T("Marketing", "Request Funding (HoM)", "HoM requests funding from Finance/SuperAdmin", "Request created (PENDING), SuperAdmin + Finance notified (NOT HoM's own team)", "P0 - Critical", "HoM")
T("Marketing", "Approve Funding Req", "HoM/Finance approves funding request", "Request APPROVED, matching funding ledger row created in same tx", "P0 - Critical", "HoM / Finance")
T("Marketing", "Reject Funding Req", "HoM/Finance rejects funding request", "Request REJECTED, requester notified", "P1 - High", "HoM / Finance")
T("Marketing", "Funding Balance", "View funding balance (COMPLETED incoming - APPROVED ad spend)", "Balance shown accurately", "P1 - High", "Media Buyer / HoM")
T("Marketing", "Funding Page Layout", "Verify two-section layout: 'Funds I Received' + 'Funds I Distribute' (HoM)", "Both sections visible with correct tabs (Transfers / My Requests / MB Requests)", "P1 - High", "HoM")
T("Marketing", "Disputed Banner", "Verify danger banner when disputed count > 0", "Banner shown with 'Review' CTA deep-linking to ?status=DISPUTED", "P2 - Medium", "HoM")

# Ad Spend
T("Marketing", "Add Expense", "MB logs daily ad spend via multi-line 'Add Expense' modal (date, N lines with campaign + product + amount + platform + screenshot)", "All lines created in one batch tx, HoM gets ONE notification", "P0 - Critical", "Media Buyer")
T("Marketing", "Ad Spend Screenshot", "Attempt ad spend without screenshot", "Server rejects — screenshot mandatory", "P0 - Critical", "Media Buyer")
T("Marketing", "Ad Spend Grouped", "View ad spend accordion grouped by (date x MB)", "Each day/MB row shows rolled-up status (PENDING/APPROVED/REJECTED/MIXED)", "P1 - High", "HoM")
T("Marketing", "Approve Ad Spend", "HoM approves individual ad spend line", "Status → APPROVED", "P1 - High", "HoM")
T("Marketing", "Reject Ad Spend", "HoM rejects ad spend line", "Status → REJECTED, MB notified", "P1 - High", "HoM")
T("Marketing", "High CPA Warning", "Ad spend vs leads exceeds threshold", "Auto-alert sent to HoM", "P2 - Medium", "System")
T("Marketing", "CPA Calculation", "CPA = Total Ad Spend / Total Orders (all statuses)", "CPA metric correct on dashboard", "P1 - High", "HoM")
T("Marketing", "True ROAS", "ROAS = Revenue from DELIVERED orders / Total Ad Spend", "ROAS metric correct (only DELIVERED revenue)", "P1 - High", "HoM")

# Cross-Funnel
T("Marketing", "Cross-Funnel Page", "View cross-funnel attempts on /admin/marketing/cross-funnel", "Only runner-up MB's attempts shown (not orders, not in pipeline/profit)", "P2 - Medium", "Media Buyer / HoM")

# Marketing Orders
T("Marketing", "My Orders (MB)", "Media Buyer views own orders from their campaigns", "Only own orders shown, not other MBs'", "P1 - High", "Media Buyer")
T("Marketing", "All Orders (HoM)", "HoM views all marketing orders across MBs", "All branch orders shown with MB attribution", "P1 - High", "HoM")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 7: CS — ALL USE CASES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_CS = "7. CUSTOMER SUCCESS (CS) — FULL WORKFLOW"

# Dispatch & Assignment
T("CS", "Manual Dispatch (Default)", "Order arrives as UNPROCESSED — HoCS manually assigns to CS agent", "Order → CS_ASSIGNED, agent sees it in queue", "P0 - Critical", "HoCS")
T("CS", "Load Balanced Dispatch", "Switch dispatch to load_balanced — new order auto-assigned to agent with fewest pending", "Order auto-assigned, tie-breaker: most idle agent", "P1 - High", "HoCS (Settings)")
T("CS", "Performance Dispatch", "Switch dispatch to performance — new order assigned to highest delivery-rate agent", "Order assigned to top-performer", "P1 - High", "HoCS (Settings)")
T("CS", "Claim Dispatch", "Switch dispatch to claim — agents race to claim orders from pool", "Agent claims order with atomic lock, CS_ENGAGED", "P1 - High", "CS Agent")
T("CS", "Claim Cap", "Agent at claim cap (default 2 unconfirmed) tries to claim", "Claim blocked server-side", "P1 - High", "CS Agent")
T("CS", "Hot Swap", "HoCS selects orders from one agent, mass-reassigns to another", "Orders reassigned, audit trail shows old/new agent", "P0 - Critical", "HoCS")
T("CS", "Bulk Assign", "HoCS selects multiple unassigned orders, assigns to one CS agent", "All orders → CS_ASSIGNED to selected agent (one atomic operation)", "P1 - High", "HoCS")

# Order Engagement
T("CS", "Engage Order", "CS agent clicks Engage on CS_ASSIGNED order", "Order → CS_ENGAGED, locked to agent for 15 min", "P0 - Critical", "CS Agent")
T("CS", "Engage Unassigned", "CS agent takes unassigned UNPROCESSED order (if dispatch allows)", "Order → CS_ENGAGED if agent has capacity", "P1 - High", "CS Agent")
T("CS", "Phone Masking", "View order detail — customer phone shown as 0803****1234", "Phone masked in DOM, network, console. Never exposed raw", "P0 - Critical", "CS Agent")

# Communication
T("CS", "Call (VOIP)", "CS agent clicks Call — VOIP connects both parties", "Call initiated via Twilio/bridge, agent never sees raw number", "P0 - Critical", "CS Agent")
T("CS", "Call (Manual)", "VOIP disabled — CS agent logs manual call", "Manual call logged, call_status = MANUAL_CALL", "P1 - High", "CS Agent")
T("CS", "Send SMS", "CS agent sends SMS from communication panel", "SMS sent via platform bridge, outbound_messages row + timeline event created", "P1 - High", "CS Agent")
T("CS", "Send WhatsApp", "CS agent sends WhatsApp template message (no freeform)", "Template placeholders auto-filled, sent via bridge, logged", "P1 - High", "CS Agent")
T("CS", "Message Templates", "HoCS creates/edits message templates (branch-scoped)", "Template saved with placeholders, available to agents", "P1 - High", "HoCS")

# Confirm Gate
T("CS", "Confirm — Call Gate", "CS agent clicks Confirm — requires call_duration >= 15s", "Confirm button disabled until qualifying call exists", "P0 - Critical", "CS Agent")
T("CS", "Confirm — Admin Bypass", "SuperAdmin/Admin confirms without call log", "Confirm allowed (admin-class bypass)", "P0 - Critical", "SuperAdmin / Admin")
T("CS", "Confirm — Branch Admin Bypass", "Branch Admin confirms order in same branch without call log", "Confirm allowed (same-branch bypass)", "P1 - High", "Branch Admin")
T("CS", "Confirm — HoCS Bypass", "HoCS confirms using any rep's call on the order", "Confirm allowed using another agent's qualifying call", "P1 - High", "HoCS")
T("CS", "Confirm — Stock Check", "Confirm fires assertGlobalAvailabilityForOrder", "If insufficient global stock → confirm blocked with error", "P0 - Critical", "CS Agent")
T("CS", "Confirm Success", "Order confirmed with passing gates", "Order → CONFIRMED, stock: Available → Reserved", "P0 - Critical", "CS Agent")

# Cancel
T("CS", "Cancel Order", "CS/HoCS cancels UNPROCESSED/CS_ASSIGNED/CS_ENGAGED order with reason (>=10 chars)", "Order → CANCELLED, reason stored, no stock effects", "P0 - Critical", "CS Agent / HoCS")
T("CS", "Cancel — Short Reason", "Cancel with reason < 10 characters", "Server rejects — minimum 10 chars required", "P1 - High", "CS Agent")

# Allocate to 3PL (Rider Proxy)
T("CS", "Share to 3PL", "CS agent allocates CONFIRMED order to 3PL location (CONFIRMED → AGENT_ASSIGNED)", "Stock: Reserved → Allocated_to_3PL at chosen location, movements logged", "P0 - Critical", "CS Agent")
T("CS", "Share to 3PL — WhatsApp", "Click 'Share to 3PL' — copies rendered template to clipboard, opens WhatsApp group", "Rendered body in clipboard, WhatsApp group opened in new tab", "P1 - High", "CS Agent")
T("CS", "Share to 3PL — Stock Gate", "Allocate to location without available stock", "Server rejects — assertLocationCanFulfillOrder fails", "P0 - Critical", "CS Agent")

# Mark Delivered (Rider Proxy)
T("CS", "Mark Delivered", "CS agent marks AGENT_ASSIGNED → DELIVERED via follow-up call", "Stock deducted (FIFO), commission triggered, revenue recognized", "P0 - Critical", "CS Agent")
T("CS", "Mark Delivered — Optional Note", "Mark delivered with delivery note + proof URL (both optional)", "If provided, stored on order; transition not blocked without them", "P0 - Critical", "CS Agent")
T("CS", "Mark Delivered — Location Required", "Attempt DELIVERED without logistics_location_id on order", "Server rejects — location required for inventory deduction", "P1 - High", "CS Agent")

# CS cannot mark REMITTED
T("CS", "REMITTED Block", "CS agent attempts to mark order REMITTED", "Action not available — REMITTED is accountant-only", "P0 - Critical", "CS Agent")

# Order Updates
T("CS", "Update Order", "CS updates address, quantity, or upsell on engaged order", "Version snapshot created (temporal table), timeline event logged", "P1 - High", "CS Agent")
T("CS", "Price Edit Request", "CS requests order line price change", "Permission request created, HoCS/Branch Admin approves", "P2 - Medium", "CS Agent")

# Queue & Monitoring
T("CS", "Order Queue", "View CS order queue with status filters", "Orders grouped/filtered by status, assigned agent shown", "P0 - Critical", "CS Agent / HoCS")
T("CS", "Cart Abandonment", "View cart abandonment tab — pending and abandoned carts", "Carts listed with phone (masked), product, timestamp", "P2 - Medium", "CS Agent / HoCS")
T("CS", "Callback Queue", "View scheduled callbacks", "Callbacks listed with scheduled time, order link", "P2 - Medium", "CS Agent")
T("CS", "Duplicate Flagging", "View flagged duplicate orders", "Duplicates shown side-by-side for merge/dismiss", "P2 - Medium", "CS Agent / HoCS")
T("CS", "CS Leaderboard", "View CS leaderboard (delivery rate, confirmation rate)", "Agents ranked by performance metrics", "P1 - High", "HoCS")
T("CS", "Inactive Agents", "View inactive CS agents", "Agents with no recent activity flagged", "P2 - Medium", "HoCS")

# Supervisor Mirror View (Socket.io)
T("CS", "Supervisor Mirror View", "HoCS opens live view of CS agent's screen state", "Read-only view of agent's current route/panel/order", "P1 - High", "HoCS")
T("CS", "Being Observed", "CS agent sees 'Being Observed' indicator when HoCS mirrors", "Indicator visible, transparency requirement met", "P1 - High", "CS Agent")

# Transfer blocked
T("CS", "Agent Transfer Block", "CS agent attempts to transfer order to another agent", "Action not available — only HoCS can reassign", "P1 - High", "CS Agent")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 8: LOGISTICS — HEAD OF LOGISTICS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_LOGISTICS = "8. LOGISTICS — HEAD OF LOGISTICS & 3PL"

T("Logistics", "Dashboard", "View logistics overview page", "Pipeline counts, health metrics, stuck orders visible", "P0 - Critical", "HoLogistics")
T("Logistics", "Orders View", "View logistics orders list (dispatched, in-transit, delivered)", "Orders filtered to logistics-relevant statuses", "P0 - Critical", "HoLogistics")
T("Logistics", "Bulk Transition", "HoLogistics bulk-transitions selected orders (e.g., DISPATCHED → IN_TRANSIT)", "All selected orders transitioned atomically", "P1 - High", "HoLogistics")
T("Logistics", "Provider Performance", "View logistics team analysis — provider delivery rate, delinquency rate", "Providers ranked by deliveryRate desc, stacked-bar breakdown", "P1 - High", "HoLogistics")
T("Logistics", "Shrinkage Alerts", "View shrinkage alerts (physical vs digital mismatch)", "Alerts listed with location, product, variance", "P1 - High", "HoLogistics")
T("Logistics", "Stuck Orders", "View orders stuck in DISPATCHED/IN_TRANSIT too long", "Stuck orders flagged with duration", "P1 - High", "HoLogistics")
T("Logistics", "Transfer Delays", "View delayed stock transfers", "Transfers past expected arrival flagged", "P2 - Medium", "HoLogistics")
T("Logistics", "Health Dashboard", "View logistics health dashboard (overall system health)", "Health metrics aggregated across all locations", "P2 - Medium", "HoLogistics")
T("Logistics", "Delivery Confirm", "Rider/3PL submits delivery confirmation (OTP/GPS/signature)", "Confirmation created (PENDING), HoLogistics reviews", "P1 - High", "TPL Rider")
T("Logistics", "Approve Delivery", "HoLogistics approves delivery confirmation", "Confirmation APPROVED, order status updated", "P1 - High", "HoLogistics")
T("Logistics", "Reject Delivery", "HoLogistics rejects delivery confirmation", "Confirmation REJECTED, rider notified", "P1 - High", "HoLogistics")
T("Logistics", "Transfer Remittance", "TPL Manager creates transfer remittance (stock return to warehouse)", "Remittance created (PENDING)", "P2 - Medium", "TPL Manager")
T("Logistics", "Mark Remittance Received", "HoLogistics marks transfer remittance received", "Remittance RECEIVED, stock updated at source", "P2 - Medium", "HoLogistics")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 9: FULL ORDER LIFECYCLE (End-to-End)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_ORDER_E2E = "9. ORDER LIFECYCLE — END-TO-END FLOWS"

T("Order E2E", "Happy Path", "Full lifecycle: Edge submit → CS_ASSIGNED → CS_ENGAGED → CONFIRMED → AGENT_ASSIGNED → DELIVERED → REMITTED", "Each transition logged in audit + timeline, stock flows correct at each step", "P0 - Critical", "Multi-Role")
T("Order E2E", "Cancel Early", "Edge submit → UNPROCESSED → CANCELLED (reason >=10 chars)", "No stock touched, reason stored, audit logged", "P0 - Critical", "CS Agent")
T("Order E2E", "Cancel After Engage", "CS_ENGAGED → CANCELLED", "No stock touched (was never reserved), reason stored", "P0 - Critical", "CS Agent")
T("Order E2E", "Partial Delivery", "DISPATCHED/IN_TRANSIT → PARTIALLY_DELIVERED (specify delivered vs returned qty)", "Delivered portion completes, returned portion enters return flow", "P1 - High", "TPL Rider")
T("Order E2E", "Return", "IN_TRANSIT → RETURNED (mandatory return reason)", "Return flow begins, clawback triggered for MB + CS", "P1 - High", "TPL Rider")
T("Order E2E", "Restock After Return", "RETURNED → RESTOCKED (3PL marks sellable, quality check)", "Stock: +1 at 3PL local inventory", "P1 - High", "TPL Manager")
T("Order E2E", "Write Off", "RETURNED → WRITTEN_OFF (mandatory damage note)", "Logged as Operational Loss in Finance", "P1 - High", "TPL Manager")
T("Order E2E", "State Skip Block", "Attempt UNPROCESSED → DISPATCHED (skip states)", "Server rejects — state machine enforced", "P0 - Critical", "Any")
T("Order E2E", "Timeline", "View order timeline — all events with actor name + timestamp", "Every state transition + update + message shown chronologically", "P0 - Critical", "All Roles")
T("Order E2E", "Offline CS Order", "CS creates order manually via createOffline", "Order created as UNPROCESSED, source marked as offline", "P1 - High", "CS Agent")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 10: FINANCE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_FINANCE = "10. FINANCE — P&L, INVOICES, REMITTANCES, PAYOUTS"

T("Finance", "P&L Report", "View profit & loss report (Revenue - Landed COGS - Ad Spend - 3PL Fee - Delivery Fee - Commission)", "True net cash profit shown; date filter applied to all cost lines", "P0 - Critical", "Finance / SuperAdmin")
T("Finance", "P&L Date Filter", "Apply date filter to P&L — all MVs scoped by date", "Ad spend by spend_date, commission by period_month, deliveries by delivery_date", "P0 - Critical", "Finance / SuperAdmin")
T("Finance", "Invoice List", "View invoices list with pagination + status filter", "Invoices listed with reference (INV-YYYY-XXXX), amount, status", "P1 - High", "Finance")
T("Finance", "Invoice Detail", "View invoice detail", "Line items, amounts, payment status shown", "P1 - High", "Finance")
T("Finance", "Invoice Overdue Flag", "System auto-flags overdue invoices", "Invoices past due date marked OVERDUE", "P2 - Medium", "System")
T("Finance", "Budget Set", "Finance sets budget per department/campaign", "Budget created with amount limit", "P1 - High", "Finance")
T("Finance", "Budget Warning", "Request exceeds remaining budget", "Warning shown (approval still possible with reason)", "P1 - High", "Finance")
T("Finance", "Approval Queue", "Finance reviews approval requests (budget overrides, etc.)", "Requests listed with approve/reject/query modal", "P1 - High", "Finance")
T("Finance", "Cash Remittance Create", "Accountant creates cash remittance from delivered orders (same logistics location)", "Remittance created, orders linked; server validates one-location-per-remittance", "P0 - Critical", "Finance")
T("Finance", "Cash Remittance Mark Received Now", "Create remittance with 'Mark received now' checked", "Remittance RECEIVED + all linked orders DELIVERED → REMITTED in one tx", "P0 - Critical", "Finance")
T("Finance", "Cash Remittance Mark Later", "Create remittance (SENT), then mark received from detail page", "DELIVERED → REMITTED cascaded on mark-received", "P1 - High", "Finance")
T("Finance", "Cash Remittance Dispute", "Dispute a cash remittance", "Status → DISPUTED, order status untouched", "P1 - High", "Finance")
T("Finance", "Payout Workspace", "Finance views payroll batches in PENDING_FINANCE/PAID", "Batches listed with per-staff payout lines, bank details visible", "P1 - High", "Finance")
T("Finance", "Payout Export", "Export payout CSV/XLSX with bank details", "File downloaded with correct bank info per staff", "P1 - High", "Finance")
T("Finance", "MV Refresh", "Materialized views refresh every 15 minutes (cron)", "CEO dashboard numbers reflect data within 15-min staleness", "P1 - High", "System")
T("Finance", "CLS Enforcement", "Non-finance user accesses order with cost fields", "cost_price, landed_cost, margin stripped from response", "P0 - Critical", "CS Agent")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 11: HR & PAYROLL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_HR = "11. HR — COMMISSION PLANS, PAYROLL, ADJUSTMENTS"

# Commission Plans
T("HR", "Create Commission Plan", "HR/Head creates commission plan for a role (base salary, thresholds, multipliers)", "Plan created, applicable to staff with matching role", "P0 - Critical", "HR Manager / HoCS / HoM / HoLog")
T("HR", "Edit Commission Plan", "Edit existing commission plan rules", "Plan updated, new rules apply to next settlement", "P1 - High", "HR Manager / HoCS")
T("HR", "Plan Scoping", "Head of CS can only create/edit CS_AGENT plans", "Other department roles not selectable", "P1 - High", "HoCS")
T("HR", "Plan Scoping — HoM", "Head of Marketing can only create/edit MEDIA_BUYER plans", "Other department roles not selectable", "P1 - High", "HoM")
T("HR", "Plan Scoping — HR", "HR Manager can create/edit plans for all roles", "All department roles available", "P1 - High", "HR Manager")

# Payroll Batches
T("HR", "Generate Batch", "Head generates payroll batch for their dept (branch x month)", "Batch created in DRAFT with all staff payouts computed", "P0 - Critical", "HoCS / HoM / HoLog")
T("HR", "Generate — Default Zero", "Staff member with no commission plan included in batch", "Staff appears with $0 payout (not skipped), HR can adjust", "P0 - Critical", "HoCS")
T("HR", "Generate — Month Default", "Generate modal pre-fills current YYYY-MM", "Month picker shows current month by default", "P2 - Medium", "HoCS")
T("HR", "Submit Batch", "Head submits DRAFT → PENDING_HR", "Batch locked from editing by head, HR notified", "P0 - Critical", "HoCS / HoM / HoLog")
T("HR", "HR Review", "HR reviews PENDING_HR batch — adds adjustments, attaches notes", "Adjustments attached to individual payouts", "P0 - Critical", "HR Manager")
T("HR", "Add Adjustment (DRAFT)", "Head adds adjustment during DRAFT stage", "Adjustment attached, payout recalculated", "P1 - High", "HoCS")
T("HR", "Add Adjustment (PENDING_HR)", "HR adds adjustment during review stage", "Adjustment attached, payout recalculated", "P1 - High", "HR Manager")
T("HR", "Approve Batch", "HR approves PENDING_HR → PENDING_FINANCE", "Batch forwarded to Finance, Finance notified", "P0 - Critical", "HR Manager")
T("HR", "Reject Batch (HR)", "HR rejects PENDING_HR → DRAFT", "Reason required (>=10 chars), head notified, batch unlocked for re-edit", "P1 - High", "HR Manager")
T("HR", "Mark Paid", "Finance marks PENDING_FINANCE → PAID (with financeReference)", "All child payouts → PAID, batch closed", "P0 - Critical", "Finance")
T("HR", "Reject Batch (Finance)", "Finance rejects PENDING_FINANCE → PENDING_HR", "Reason required, HR notified", "P1 - High", "Finance")
T("HR", "Re-generate DRAFT", "Head re-generates an existing DRAFT batch", "Old payouts wiped, fresh computation from latest data", "P1 - High", "HoCS")
T("HR", "Re-generate Non-DRAFT", "Attempt to re-generate PENDING_HR batch", "Server rejects — must reject first", "P1 - High", "HoCS")

# Adjustments
T("HR", "Bonus Adjustment", "HR creates bonus adjustment (Special Service / Extra Shift / Performance)", "Adjustment created, requires admin approval, distinct line in payout", "P1 - High", "HR Manager")
T("HR", "Clawback", "Order returned → system auto-creates PENDING_DEDUCTION for MB + CS", "Negative line item in next payout for both", "P0 - Critical", "System")

# Commission Rules
T("HR", "Commission by Delivered Date", "Commission calculated on DELIVERED_AT, not CREATED_AT", "January order delivered in February → February payout", "P0 - Critical", "System")

# Onboarding
T("HR", "Staff Onboarding", "New staff completes onboarding documents", "Documents submitted, HR reviews and approves", "P1 - High", "New Staff / HR")
T("HR", "Onboarding Request Changes", "HR requests changes on submitted documents", "Status → PENDING_CHANGES, staff re-edits", "P2 - Medium", "HR Manager")

# Settlement
T("HR", "Settlement — Monthly Only", "Settlement window is monthly (no weekly/biweekly)", "UI only shows MONTHLY option", "P1 - High", "HR Manager")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 12: USER MANAGEMENT & RBAC
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_USERS = "12. USER MANAGEMENT, RBAC & PERMISSIONS"

T("Users", "Create Staff", "Admin/HR creates new staff user with role + branch", "User created, assigned to branch, permissions stamped from template", "P0 - Critical", "Admin / HR Manager")
T("Users", "Create — Head Warning", "Create user with a Head role when another holder exists", "Inline warning: 'X already holds {role}' + confirm modal", "P1 - High", "Admin")
T("Users", "Create — Admin Approval", "Admin creates another ADMIN → permission request flow", "Permission request created for SuperAdmin approval", "P0 - Critical", "Admin")
T("Users", "Edit Staff", "Update user details (name, email, role, branch)", "Changes saved, audit trail captured", "P1 - High", "Admin / HR Manager")
T("Users", "Deactivate Staff", "Deactivate a staff user", "User deactivated, cannot login, session killed", "P1 - High", "Admin / HR Manager")
T("Users", "Finance Hat Assign", "Assign Finance hat to a non-FINANCE_OFFICER user", "Old holder displaced (notification sent), new holder gets hat (notification sent)", "P0 - Critical", "Admin")
T("Users", "Finance Hat Singleton", "Attempt to set Finance hat on two users simultaneously", "Atomic swap — old cleared, new set in one tx", "P1 - High", "Admin")
T("Users", "Bank Details", "Add/edit payout bank details (bank name, account name, account number)", "Bank details saved, visible only in finance flows", "P1 - High", "Staff (self) / HR")
T("Users", "Role Templates", "View/create/edit role templates with permission codes", "Templates assignable to users, permissions inherited", "P1 - High", "Admin / HR")
T("Users", "Permission Backfill", "Run permission stamp backfill after catalog changes", "Existing users pick up new permission codes", "P2 - Medium", "Admin (CLI)")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 13: MIRROR MODE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_MIRROR = "13. MIRROR MODE (FULL-SESSION IMPERSONATION)"

T("Mirror", "Start Mirror", "Admin clicks 'Mirror user' on user detail page", "App renders as target user — their role, branch, sidebar, theme", "P0 - Critical", "SuperAdmin / Admin")
T("Mirror", "Read-Only", "While mirroring, attempt any tRPC mutation", "Mutation blocked with 'Read-only while mirroring' error", "P0 - Critical", "SuperAdmin / Admin")
T("Mirror", "Visual Chrome", "While mirroring — green border, 'Exit mirror' pill in header", "4px green border on viewport, pulsing exit pill visible", "P1 - High", "SuperAdmin / Admin")
T("Mirror", "Exit Mirror", "Click 'Exit mirror' pill", "Original session restored, redirect to /admin", "P0 - Critical", "SuperAdmin / Admin")
T("Mirror", "No Notifications Mutation", "While mirroring, click notification bell", "Notifications render but markAsRead is a no-op", "P1 - High", "SuperAdmin / Admin")
T("Mirror", "No Socket Broadcast", "While mirroring, navigate — no agent:state_update emitted", "Target user's lastActionAt unchanged", "P1 - High", "SuperAdmin / Admin")
T("Mirror", "Nested Block", "While mirroring User A, attempt to mirror User B", "Mirror button disabled — must exit first", "P1 - High", "SuperAdmin / Admin")
T("Mirror", "Permission Matrix", "HoCS mirrors CS_AGENT (allowed); HoCS mirrors MEDIA_BUYER (blocked)", "Role matrix enforced per canMirror()", "P1 - High", "HoCS")
T("Mirror", "Phone Masking in Mirror", "Admin mirrors CS agent — customer phones still masked", "No side-channel for raw PII via mirror", "P0 - Critical", "Admin")
T("Mirror", "Audit Trail", "Mirror session logged in mirror_sessions table", "actor_id, target_id, started_at, ended_at all recorded", "P0 - Critical", "System")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 14: PERMISSION REQUESTS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_PERM_REQ = "14. PERMISSION REQUESTS & APPROVAL WORKFLOWS"

T("Perm Requests", "User Creation Request", "Non-admin submits user creation → PENDING for HR approval", "Request created, HR notified", "P1 - High", "HoCS / HoM")
T("Perm Requests", "Role Change Request", "Non-admin requests role change → PENDING for HR", "Request created, HR notified", "P1 - High", "Any Role")
T("Perm Requests", "Product Archive Request", "Stock Manager requests product archive → PENDING for SuperAdmin ONLY", "Request created, SuperAdmin notified (not HR, not Admin)", "P1 - High", "Stock Manager")
T("Perm Requests", "Order Line Price Change", "CS requests line price edit → PENDING for HoCS/Branch Admin/HoLogistics", "Request created, approvers notified", "P2 - Medium", "CS Agent")
T("Perm Requests", "Order Deletion Request", "CS requests order deletion → PENDING for approvers", "Request created, approvers notified", "P2 - Medium", "CS Agent")
T("Perm Requests", "Approve Request", "Approver approves pending request", "Action executed (user created / role changed / product archived)", "P1 - High", "Approver")
T("Perm Requests", "Reject Request", "Approver rejects pending request", "Request REJECTED, requester notified", "P1 - High", "Approver")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 15: BRANCHES
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_BRANCHES = "15. BRANCH MANAGEMENT"

T("Branches", "Create Branch", "SuperAdmin creates a new branch", "Branch created, available for user assignment", "P0 - Critical", "SuperAdmin")
T("Branches", "Edit Branch", "Update branch name/details", "Changes saved", "P1 - High", "SuperAdmin")
T("Branches", "Assign User", "Add user to a branch", "User can switch to that branch, sees branch-scoped data", "P0 - Critical", "Admin / Branch Admin")
T("Branches", "Remove User", "Remove user from a branch", "User loses access to that branch's data", "P1 - High", "Admin / Branch Admin")
T("Branches", "Switch Branch", "User switches active branch via sidebar", "Session updated, all data reloads for new branch context", "P0 - Critical", "Multi-Branch User")
T("Branches", "Branch Scoping", "Orders, campaigns, funding etc. filtered by active branch", "Only current branch data shown; SuperAdmin sees all", "P0 - Critical", "All Roles")
T("Branches", "Org-Wide Heads", "HoCS/HoM/HoLogistics see all branches (currentBranchId=NULL)", "Cross-branch visibility for domain data", "P1 - High", "HoCS / HoM / HoLog")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 16: DASHBOARDS & REPORTING
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_DASHBOARDS = "16. DASHBOARDS & REPORTING"

T("Dashboard", "Admin Landing", "SuperAdmin/Admin lands on /admin — lightweight quick overview", "Today's status counts, active orders, pending approvals; loads < 200ms", "P0 - Critical", "SuperAdmin / Admin")
T("Dashboard", "CEO Executive", "Navigate to /admin/ceo — full Executive Overview", "Revenue, profit, pipeline charts, leaderboards, branch breakdown", "P0 - Critical", "SuperAdmin / Admin")
T("Dashboard", "CS Dashboard", "CS Agent lands on /admin — CS-specific dashboard", "Own orders, queue, performance metrics", "P0 - Critical", "CS Agent")
T("Dashboard", "Marketing Dashboard", "Media Buyer lands on /admin — marketing dashboard", "Own campaigns, ad spend, funding, orders", "P0 - Critical", "Media Buyer")
T("Dashboard", "Finance Dashboard", "Finance Officer lands on /admin — finance dashboard", "P&L, invoices, remittances, payroll", "P0 - Critical", "Finance Officer")
T("Dashboard", "HR Dashboard", "HR Manager views HR pages", "Payroll, commission plans, adjustments, onboarding", "P0 - Critical", "HR Manager")
T("Dashboard", "Logistics Dashboard", "HoLogistics lands on /admin — logistics dashboard", "Pipeline, health, providers, team analysis", "P0 - Critical", "HoLogistics")
T("Dashboard", "Branch Admin", "Branch Admin lands on /admin — branch-scoped view", "Own branch users, settings, reports", "P1 - High", "Branch Admin")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 17: AUDIT & COMPLIANCE
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_AUDIT = "17. AUDIT TRAIL & COMPLIANCE"

T("Audit", "Global Log", "View global audit log (/admin/audit)", "All mutations listed with actor, action, old/new values, timestamp", "P0 - Critical", "SuperAdmin / Admin / Finance")
T("Audit", "Table Filter", "Filter audit log by table (orders, products, users, etc.)", "Results filtered to selected table", "P1 - High", "SuperAdmin")
T("Audit", "Actor Filter", "Filter audit log by actor (who made the change)", "Results filtered to selected user", "P1 - High", "SuperAdmin")
T("Audit", "Time Travel", "Query state of any record at a specific point in time", "Historical snapshot shown accurately", "P1 - High", "SuperAdmin")
T("Audit", "Record History", "View audit trail for a specific record (e.g., order)", "All changes to that record shown chronologically", "P1 - High", "All Roles (scoped)")
T("Audit", "Mirror Sessions", "View Mirror Mode sessions card above audit table", "Active sessions show pulsing 'Active' badge; closed show ended_at", "P1 - High", "SuperAdmin / Admin")
T("Audit", "Export CSV", "Export audit log as CSV", "CSV downloaded with all columns", "P2 - Medium", "SuperAdmin")
T("Audit", "Actor Attribution", "Every write shows correct actor (never 'System' for user actions)", "withActor() ensures actor_id set in every tx", "P0 - Critical", "SuperAdmin")
T("Audit", "Immutability", "Attempt to modify/delete an audit entry", "Blocked — audit trail is permanent and immutable", "P0 - Critical", "SuperAdmin")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 18: NOTIFICATIONS
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_NOTIFICATIONS = "18. NOTIFICATIONS — IN-APP, PUSH, BROADCAST"

T("Notifications", "In-App Bell", "Click notification bell — list of notifications shown", "Notifications listed with type-colored dots, newest first", "P0 - Critical", "All Roles")
T("Notifications", "Mark Read", "Mark individual notification as read", "Dot removed, unread count decremented", "P1 - High", "All Roles")
T("Notifications", "Mark All Read", "Click 'Mark all as read'", "All notifications marked read, count → 0", "P1 - High", "All Roles")
T("Notifications", "Click-Through", "Click notification → navigate to relevant page", "Deep-link to order/payroll/funding etc.", "P1 - High", "All Roles")
T("Notifications", "Push Subscribe", "Grant browser push permission", "VAPID subscription saved, push enabled on device", "P1 - High", "All Roles")
T("Notifications", "Push Receive", "Trigger event that generates push (e.g., order assigned)", "Push notification shown on device lock screen / notification center", "P0 - Critical", "All Roles")
T("Notifications", "Push Click", "Click push notification on device", "App opens to relevant page (data.url), ack POSTed with 'clicked'", "P1 - High", "All Roles")
T("Notifications", "iOS Install Gate", "iOS user without PWA installed to Home Screen", "Install banner shown — push requires Home Screen add", "P2 - Medium", "iOS Users")
T("Notifications", "Broadcast Push", "Admin broadcasts push to ALL / ROLE / USER target", "Push delivered to all matching subscriptions, delivery log created", "P1 - High", "Admin / Branch Admin")
T("Notifications", "Broadcast Scope", "HoCS broadcasts — only CS_AGENT targets allowed", "Out-of-scope targets rejected server-side", "P1 - High", "HoCS")
T("Notifications", "Automation CRON", "Create CRON automation rule (e.g., daily stuck-order reminder)", "Cron job registered, fires at scheduled time", "P2 - Medium", "Admin")
T("Notifications", "Automation Toggle", "Disable an active automation rule", "Cron job unregistered; re-enable re-registers", "P2 - Medium", "Admin")
T("Notifications", "Delivery Log", "View push delivery log (SENT/SHOWN/CLICKED/FAILED)", "All push attempts listed with status + timestamps", "P2 - Medium", "Admin")
T("Notifications", "Resend Failed", "Resend a FAILED push notification", "New push_delivery_log row created, push re-sent", "P2 - Medium", "Admin")
T("Notifications", "Non-Blocking", "Mutation completes without waiting for notification fan-out", "Notification enqueued via enqueueCreate, response not delayed", "P0 - Critical", "System")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 19: SETTINGS & SYSTEM CONFIG
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_SETTINGS = "19. SETTINGS & SYSTEM CONFIGURATION"

T("Settings", "Dispatch Mode", "Change CS dispatch mode (manual/load_balanced/performance/claim)", "Mode saved, next orders use new dispatch strategy", "P1 - High", "SuperAdmin")
T("Settings", "VOIP Toggle", "Enable/disable VOIP", "Toggle validates provider config before enabling", "P1 - High", "SuperAdmin")
T("Settings", "Default Theme", "Set org-wide default theme", "New users and users with null theme get this default", "P2 - Medium", "SuperAdmin")
T("Settings", "Client UI Config", "Update client UI branding/colors", "Changes reflected across all users on next load", "P2 - Medium", "SuperAdmin")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 20: THEME & ACCESSIBILITY
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_THEME = "20. THEME, FONT SCALE & ACCESSIBILITY"

T("Theme", "Theme Switch", "Switch between 6 themes (system/light/dark/dim/ink/soft)", "Theme applied immediately, no flash on page reload", "P1 - High", "All Roles")
T("Theme", "Theme Persistence", "Change theme → logout → login on different device", "Same theme applied (synced via server)", "P1 - High", "All Roles")
T("Theme", "No Flash", "Hard-refresh page", "Theme applied before first paint (inline boot script)", "P2 - Medium", "All Roles")
T("Font Scale", "Scale Switch", "Switch font scale (base/large/xlarge)", "Root font-size changes, all UI scales proportionally", "P1 - High", "All Roles")
T("Font Scale", "Scale Persistence", "Change scale → login on different device", "Same scale applied (synced via server)", "P1 - High", "All Roles")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 21: 3PL PARTNER PORTAL
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_TPL = "21. 3PL PARTNER PORTAL (TPL DASHBOARD)"

T("TPL Portal", "Login", "TPL Manager logs in — sees simplified TPL dashboard", "Only TPL-relevant data visible (own location orders, stock, remittances)", "P0 - Critical", "TPL Manager")
T("TPL Portal", "Inventory View", "View stock at own 3PL location", "Only own location inventory shown", "P1 - High", "TPL Manager")
T("TPL Portal", "Orders View", "View orders allocated to own location", "Only orders at own location shown", "P1 - High", "TPL Manager")
T("TPL Portal", "Verify Transfer", "Receive and verify incoming stock transfer", "Actual received qty recorded, inventory updated", "P0 - Critical", "TPL Manager")
T("TPL Portal", "Remittance", "Create transfer remittance (return cash to HQ)", "Remittance created with receipt upload", "P1 - High", "TPL Manager")
T("TPL Portal", "Notifications", "View notifications in TPL dashboard", "Only TPL-relevant notifications shown", "P2 - Medium", "TPL Manager")
T("TPL Portal", "Settings", "View/update TPL settings", "Settings scoped to TPL role", "P2 - Medium", "TPL Manager")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 22: RIDER (MOBILE PWA)
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_RIDER = "22. RIDER MOBILE PWA"

T("Rider", "Login", "Rider logs in on mobile — sees rider dashboard", "Mobile-optimized layout, only own deliveries", "P0 - Critical", "TPL Rider")
T("Rider", "Pickup Tab", "View Pickup tab (DISPATCHED orders)", "Orders ready for pickup listed", "P1 - High", "TPL Rider")
T("Rider", "In Transit Tab", "View In Transit tab", "Orders currently being delivered listed", "P1 - High", "TPL Rider")
T("Rider", "Start Delivery", "Click 'Start Delivery' on dispatched order", "Order → IN_TRANSIT, GPS logged", "P1 - High", "TPL Rider")
T("Rider", "Confirm Delivery", "Rider confirms delivery (OTP/signature/GPS)", "Order → DELIVERED, stock deducted", "P0 - Critical", "TPL Rider")
T("Rider", "Mark Returned", "Rider marks order as returned (mandatory reason)", "Return flow triggered, reason stored", "P1 - High", "TPL Rider")
T("Rider", "Partial Delivery", "Rider marks partial delivery (qty delivered vs returned)", "Split: delivered portion completes, returned enters return flow", "P1 - High", "TPL Rider")
T("Rider", "Offline Mode", "Rider goes offline — confirms delivery while offline", "Confirmation stored in IndexedDB, syncs within 30s of reconnect", "P1 - High", "TPL Rider")
T("Rider", "Phone Masking", "Rider views customer phone on delivery", "Phone masked (0803****1234), never raw", "P0 - Critical", "TPL Rider")

# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# SECTION 23: GLOBAL SEARCH & NAVIGATION
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
SECTION_GLOBAL = "23. GLOBAL SEARCH, NAVIGATION & UX"

T("Global", "Cmd+K Search", "Press Cmd+K — global search modal opens", "Search across orders/products/users with debounce", "P1 - High", "All Roles")
T("Global", "Search Results", "Type search query — results with type badges", "Results grouped by type, keyboard navigation works", "P1 - High", "All Roles")
T("Global", "Sidebar Navigation", "Navigate via collapsible sidebar", "All role-appropriate menu items visible, active state correct", "P0 - Critical", "All Roles")
T("Global", "Branch Switcher", "Switch branch via sidebar selector", "Data reloads for new branch context", "P0 - Critical", "Multi-Branch User")
T("Global", "Breadcrumbs", "Navigate nested pages — breadcrumb trail shown", "Breadcrumbs reflect current path, clickable parents", "P2 - Medium", "All Roles")
T("Global", "Responsive Design", "Use app on mobile viewport", "Layout adapts, tables show mobile card view", "P1 - High", "All Roles")
T("Global", "Loading Indicators", "Filter/paginate on any list page", "NavProgressBar shows at top + table overlay while loading", "P1 - High", "All Roles")
T("Global", "Error Boundaries", "Navigate to non-existent route", "404 page shown with 'Go Home' button", "P2 - Medium", "All Roles")
T("Global", "CSV Export", "Export data from orders/finance/HR/inventory pages", "CSV file downloaded with correct data", "P1 - High", "Admin / Finance / HR")

# ════════════════════════════════════════════════════════════════════════
# BUILD THE SHEETS
# ════════════════════════════════════════════════════════════════════════

# Group test cases by section
sections = [
    (SECTION_AUTH, "Auth"),
    (SECTION_PRODUCTS, "Products"),
    (SECTION_WAREHOUSES, "Warehouses"),
    (SECTION_SHIPMENTS, "Shipments"),
    (SECTION_STOCK, "Stock & Inventory"),
    (SECTION_MARKETING, "Marketing"),
    (SECTION_CS, "CS"),
    (SECTION_LOGISTICS, "Logistics"),
    (SECTION_ORDER_E2E, "Order E2E"),
    (SECTION_FINANCE, "Finance"),
    (SECTION_HR, "HR & Payroll"),
    (SECTION_USERS, "Users & RBAC"),
    (SECTION_MIRROR, "Mirror Mode"),
    (SECTION_PERM_REQ, "Perm Requests"),
    (SECTION_BRANCHES, "Branches"),
    (SECTION_DASHBOARDS, "Dashboards"),
    (SECTION_AUDIT, "Audit"),
    (SECTION_NOTIFICATIONS, "Notifications"),
    (SECTION_SETTINGS, "Settings"),
    (SECTION_THEME, "Theme & Font"),
    (SECTION_TPL, "3PL Portal"),
    (SECTION_RIDER, "Rider PWA"),
    (SECTION_GLOBAL, "Global UX"),
]

# Module to section mapping
module_section_map = {}
for section_title, _ in sections:
    module_section_map[section_title] = []

# Assign test cases to sections by order (they were added sequentially matching sections)
section_idx = 0
section_boundaries = []
prev_module = None
for i, (tc_id, module, sub, case, expected, priority, role) in enumerate(test_cases):
    if not section_boundaries or module != prev_module:
        section_boundaries.append((module, i))
    prev_module = module

# ── Sheet 1: Master (all test cases on one sheet) ──
ws = wb.active
dv = setup_sheet(ws, "UAT Master")

row = 2
current_section_idx = 0
tc_by_section_start = {}

# Build a mapping of test case modules to their section
module_to_section = {}
for tc_id, module, sub, case, expected, priority, role in test_cases:
    if module not in module_to_section:
        # Find the section for this module
        for s_title, s_short in sections:
            if module.lower() in s_title.lower() or module.lower() in s_short.lower():
                module_to_section[module] = s_title
                break
        if module not in module_to_section:
            module_to_section[module] = None

# Write all test cases grouped with section headers
written_sections = set()
last_module = None
for tc_id, module, sub, case, expected, priority, role in test_cases:
    # Check if we need a new section header
    if module != last_module:
        # Find the section for this module
        for s_title, s_short in sections:
            if s_title not in written_sections:
                # Check if this module belongs to this section
                matched = False
                for t in test_cases:
                    if t[1] == module:
                        matched = True
                        break
                if matched and (
                    module.lower() in s_short.lower()
                    or module.lower() in s_title.lower()
                    or s_short.lower() in module.lower()
                ):
                    row = add_section(ws, row, s_title)
                    written_sections.add(s_title)
                    break
        last_module = module

    row = add_row(ws, row, dv, tc_id, module, sub, case, expected, priority, role)

# ── Sheet 2: Summary Dashboard ──
ws2 = wb.create_sheet("Summary Dashboard")
ws2.sheet_properties.tabColor = "059669"

ws2.column_dimensions["A"].width = 30
ws2.column_dimensions["B"].width = 12
ws2.column_dimensions["C"].width = 12
ws2.column_dimensions["D"].width = 12
ws2.column_dimensions["E"].width = 12
ws2.column_dimensions["F"].width = 12
ws2.column_dimensions["G"].width = 12

# Summary header
ws2.merge_cells("A1:G1")
title_cell = ws2.cell(row=1, column=1, value="YANNIS EOSE — UAT TEST PLAN SUMMARY")
title_cell.font = Font(name="Inter", bold=True, size=14, color="FFFFFF")
title_cell.fill = PatternFill("solid", fgColor="111827")
title_cell.alignment = Alignment(horizontal="center", vertical="center")
for c in range(2, 8):
    ws2.cell(row=1, column=c).fill = PatternFill("solid", fgColor="111827")

# Metadata
ws2.cell(row=2, column=1, value="Generated:").font = BOLD_FONT
ws2.cell(row=2, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M")).font = BODY_FONT
ws2.cell(row=3, column=1, value="Total Test Cases:").font = BOLD_FONT
ws2.cell(row=3, column=2, value=len(test_cases)).font = BODY_FONT
ws2.cell(row=4, column=1, value="Project:").font = BOLD_FONT
ws2.cell(row=4, column=2, value="Yannis EOSE v1.0").font = BODY_FONT

# Priority breakdown
ws2.cell(row=6, column=1, value="Priority Breakdown").font = Font(name="Inter", bold=True, size=12)
ws2.cell(row=7, column=1, value="Priority").font = HEADER_FONT
ws2.cell(row=7, column=1).fill = HEADER_FILL
ws2.cell(row=7, column=2, value="Count").font = HEADER_FONT
ws2.cell(row=7, column=2).fill = HEADER_FILL

p0 = sum(1 for t in test_cases if "P0" in t[5])
p1 = sum(1 for t in test_cases if "P1" in t[5])
p2 = sum(1 for t in test_cases if "P2" in t[5])

ws2.cell(row=8, column=1, value="P0 - Critical").font = BODY_FONT
ws2.cell(row=8, column=2, value=p0).font = BODY_FONT
ws2.cell(row=9, column=1, value="P1 - High").font = BODY_FONT
ws2.cell(row=9, column=2, value=p1).font = BODY_FONT
ws2.cell(row=10, column=1, value="P2 - Medium").font = BODY_FONT
ws2.cell(row=10, column=2, value=p2).font = BODY_FONT
ws2.cell(row=11, column=1, value="Total").font = BOLD_FONT
ws2.cell(row=11, column=2, value=len(test_cases)).font = BOLD_FONT

# Module breakdown
ws2.cell(row=13, column=1, value="Module Breakdown").font = Font(name="Inter", bold=True, size=12)
headers = ["Module", "Total", "P0", "P1", "P2"]
for i, h in enumerate(headers, 1):
    cell = ws2.cell(row=14, column=i, value=h)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL

modules_order = []
for t in test_cases:
    if t[1] not in modules_order:
        modules_order.append(t[1])

row = 15
for mod in modules_order:
    mod_cases = [t for t in test_cases if t[1] == mod]
    ws2.cell(row=row, column=1, value=mod).font = BODY_FONT
    ws2.cell(row=row, column=2, value=len(mod_cases)).font = BODY_FONT
    ws2.cell(row=row, column=3, value=sum(1 for t in mod_cases if "P0" in t[5])).font = BODY_FONT
    ws2.cell(row=row, column=4, value=sum(1 for t in mod_cases if "P1" in t[5])).font = BODY_FONT
    ws2.cell(row=row, column=5, value=sum(1 for t in mod_cases if "P2" in t[5])).font = BODY_FONT
    row += 1

# Status legend
row += 2
ws2.cell(row=row, column=1, value="Status Legend").font = Font(name="Inter", bold=True, size=12)
row += 1
statuses = [
    ("PASS", PASS_FILL, "Test case passed — feature works as expected"),
    ("FAIL", FAIL_FILL, "Test case failed — defect found, log Defect ID"),
    ("PENDING", PENDING_FILL, "Test case in progress or awaiting prerequisite"),
    ("BLOCKED", BLOCKED_FILL, "Test case blocked by external dependency (env, config, etc.)"),
    ("NOT TESTED", NOT_TESTED_FILL, "Test case not yet executed"),
]
for status, fill, desc in statuses:
    ws2.cell(row=row, column=1, value=status).font = BOLD_FONT
    ws2.cell(row=row, column=1).fill = fill
    ws2.cell(row=row, column=2, value=desc).font = BODY_FONT
    ws2.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
    row += 1

# Instructions
row += 2
ws2.cell(row=row, column=1, value="How to Use This Sheet").font = Font(name="Inter", bold=True, size=12)
row += 1
instructions = [
    "1. Go to the 'UAT Master' sheet — all test cases are there in priority order.",
    "2. Column H (Status) has a dropdown: PASS / FAIL / PENDING / BLOCKED / NOT TESTED.",
    "3. Fill 'Actual Result / Notes' (col I) for FAIL cases — describe what happened.",
    "4. Fill 'Defect ID' (col L) for FAIL cases — link to your bug tracker.",
    "5. Fill 'Tested By' (col J) and 'Date Tested' (col K) for accountability.",
    "6. Filter by Status column to see all FAIL / NOT TESTED items quickly.",
    "7. Priority order: test P0 (Critical) first, then P1 (High), then P2 (Medium).",
    "8. Test by section top-to-bottom: Products → Warehouses → Shipments → Stock → Marketing → CS → Logistics → Finance → HR.",
]
for inst in instructions:
    ws2.cell(row=row, column=1, value=inst).font = BODY_FONT
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
    row += 1

# ── Save ──
OUTPUT = "/Users/Apple/Desktop/PROJECTS/ROGUE-DEVTECH/yannis-eose/Yannis_EOSE_UAT_Test_Plan.xlsx"
wb.save(OUTPUT)
print(f"Created {OUTPUT}")
print(f"Total test cases: {len(test_cases)}")
print(f"P0 Critical: {p0}")
print(f"P1 High: {p1}")
print(f"P2 Medium: {p2}")
