#!/usr/bin/env python3
"""
Generate a large, edge-case-rich orders import test workbook.

Mirrors the exact validation contract in
  apps/api/src/orders/bulk-import.service.ts  (mapRow / parseRowStatus)
  packages/shared/src/validators/orders.ts    (importOrderSchema)
  apps/web/app/features/orders/BulkImportPage.tsx (autoMapHeaders)

Output workbook has 3 sheets:
  Orders      - the data to import (headers match the CEO's template exactly)
  Reference   - the original template's rules sheet
  Test Cases  - what each row is testing + expected outcome

Usage:
  python3 scripts/gen-import-test-sheet.py --rows 1000 --out ~/Desktop/yannis-orders-import-TEST.xlsx

IMPORTANT: set VALID_PRODUCT_CODES / VALID_MB_CODES / VALID_CS_CODES /
VALID_CURRENCIES below to codes that actually exist in the target environment,
otherwise the "good" rows will fail for the wrong reason.
"""

import argparse
import random
from datetime import datetime, timedelta

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# EDIT THESE to match the target environment (Products page / Users page).
# ─────────────────────────────────────────────────────────────────────────────
VALID_PRODUCT_CODES = ["PDT-1", "PDT-2", "PDT-3", "PDT-4", "PDT-5"]

# Code -> the display name that BELONGS to that code. The importer ignores the
# name columns (only the ID drives attribution), but keeping them consistent is
# what lets you eyeball, after import, whether an order landed on the right
# person. Set these to the real names shown on your Users page.
MB_BY_CODE = {
    "USR-5": "Exre Aluya",
    "USR-6": "Tobi Aderinto",
    "USR-7": "Zainab Yusuf",
}
CS_BY_CODE = {
    "USR-12": "Annual Effiong",
    "USR-13": "Peace Nwachukwu",
    "USR-14": "Idris Bamidele",
}
VALID_MB_CODES = list(MB_BY_CODE)
VALID_CS_CODES = list(CS_BY_CODE)
VALID_CURRENCIES = ["NGN", "Nigeria", ""]  # "" = let the importer derive it

# Status labels parseRowStatus() understands (label -> resulting status).
GOOD_STATUSES = [
    "Delivered and Cash Remitted",  # REMITTED
    "Pending",                      # CS_ASSIGNED
    "No Response",                  # CS_ENGAGED
    "Rescheduled",                  # CS_ENGAGED
    "Confirmed",                    # CONFIRMED
    "Delivered",                    # DELIVERED
    "Returned",                     # RETURNED
    "Cancelled",                    # DELETED (CANCELLED is legacy-only)
    "Remitted",                     # REMITTED
    "Unprocessed",                  # UNPROCESSED
]

HEADERS = [
    "Order ID", "Date", "Name", "Phone Number", "WhatsApp Number", "Email",
    "Address", "State", "Product ID", "Product Name", "Quantity", "Cost",
    "Currency", "Gender", "Delivery Time", "More details", "Status",
    "Media-Buyer", "Media Buyer ID", "CS", "CS ID", "Delivery agent",
    "Comment 1", "Comment 2", "Comment 3",
]

FIRST = ["Chuks", "Adamu", "Ngozi", "Bola", "Emeka", "Fatima", "Tunde", "Aisha",
         "Ifeanyi", "Zainab", "Segun", "Chioma", "Musa", "Blessing", "Yusuf",
         "Amaka", "Kelechi", "Halima", "Obinna", "Funke", "Sadiq", "Uche"]
LAST = ["David", "Garba", "Okafor", "Adeyemi", "Nwosu", "Bello", "Ibrahim",
        "Eze", "Lawal", "Okonkwo", "Danjuma", "Balogun", "Umeh", "Sani",
        "Adeleke", "Chukwu", "Mohammed", "Oyelaran", "Nnamdi", "Abubakar"]
STATES = ["Lagos", "Abuja", "Rivers", "Kano", "Oyo", "Enugu", "Kaduna", "Delta",
          "Anambra", "Ogun", "Edo", "Plateau", "Borno", "Imo", "Ondo"]
STREETS = ["Adeola Odeku", "Awolowo Road", "Herbert Macaulay Way", "Aminu Kano Cres",
           "Ikorodu Road", "Allen Avenue", "Wuse Zone 4", "GRA Phase 2",
           "Opebi Link Road", "Ademola Adetokunbo"]
PRODUCT_NAMES = ["Sample Product One", "Sample Product Two", "Slim Tea Pack",
                 "Herbal Capsule", "Body Oil 250ml"]
AGENTS = ["Fomac Lagos", "Speedaf Abuja", "GIG Logistics", "Rider - Musa", ""]
DELIVERY_TIMES = ["Tomorrow", "3 Days", "Today", "Next Week", "2 Days", ""]
GENDERS = ["Male", "Female", "male", "FEMALE", ""]


def phone(rng):
    return f"0{rng.choice(['70','80','81','90','91'])}{rng.randint(10**7, 10**8 - 1)}"


def make_good_row(rng, i, base_date):
    """A clean row that should import successfully."""
    fn, ln = rng.choice(FIRST), rng.choice(LAST)
    d = base_date + timedelta(days=rng.randint(0, 240),
                              hours=rng.randint(0, 23),
                              minutes=rng.randint(0, 59))
    pidx = rng.randrange(len(VALID_PRODUCT_CODES))
    qty = rng.choice([1, 1, 1, 2, 2, 3, 5])
    unit = rng.choice([15000, 22500, 35000, 47000, 100000, 7500])
    ph = phone(rng)
    mb_code = rng.choice(VALID_MB_CODES)
    cs_code = rng.choice(VALID_CS_CODES)
    return {
        "Order ID": f"CRM-{i}",
        "Date": d.strftime("%-m/%-d/%Y"),
        "Name": f"{fn} {ln}",
        "Phone Number": ph,
        "WhatsApp Number": ph if rng.random() < 0.8 else "",
        "Email": f"{fn.lower()}.{ln.lower()}{i}@example.com" if rng.random() < 0.6 else "",
        "Address": f"{rng.randint(1, 220)} {rng.choice(STREETS)}",
        "State": rng.choice(STATES),
        "Product ID": VALID_PRODUCT_CODES[pidx],
        "Product Name": PRODUCT_NAMES[pidx % len(PRODUCT_NAMES)],
        "Quantity": qty,
        "Cost": unit * qty,
        "Currency": rng.choice(VALID_CURRENCIES),
        "Gender": rng.choice(GENDERS),
        "Delivery Time": rng.choice(DELIVERY_TIMES),
        "More details": rng.choice(["", "", "Gate is blue", "Call before arrival",
                                    "Leave with security"]),
        "Status": rng.choice(GOOD_STATUSES),
        "Media-Buyer": MB_BY_CODE[mb_code],
        "Media Buyer ID": mb_code,
        "CS": CS_BY_CODE[cs_code],
        "CS ID": cs_code,
        "Delivery agent": rng.choice(AGENTS),
        "Comment 1": rng.choice(["", "", "Customer wants morning delivery"]),
        "Comment 2": "",
        "Comment 3": "",
    }


# ─────────────────────────────────────────────────────────────────────────────
# Edge cases. Each mutator takes a clean row and breaks/bends it in one way.
# expect: FAIL = row rejected, IMPORT = row imports (possibly coerced/odd).
# ─────────────────────────────────────────────────────────────────────────────

def _mut(name, expect, reason, fn):
    return {"name": name, "expect": expect, "reason": reason, "fn": fn}


def build_mutators():
    M = []
    A = M.append

    # ── HARD FAILURES: the importer must reject these ────────────────────────
    A(_mut("Blank Order ID", "FAIL",
           'Missing external id (column "Order ID") - idempotency key is required',
           lambda r, i: r.update({"Order ID": ""})))
    A(_mut("Whitespace-only Order ID", "FAIL",
           "Trimmed to empty -> missing external id",
           lambda r, i: r.update({"Order ID": "     "})))
    A(_mut("Blank Name", "FAIL", "Missing customer name",
           lambda r, i: r.update({"Name": ""})))
    A(_mut("Name 1 char", "FAIL", "importOrderSchema requires name min 2 chars",
           lambda r, i: r.update({"Name": "A"})))
    A(_mut("Whitespace-only Name", "FAIL", "Trimmed to empty -> missing customer name",
           lambda r, i: r.update({"Name": "   "})))
    A(_mut("Blank Phone", "FAIL", "Missing customer phone",
           lambda r, i: r.update({"Phone Number": "", "WhatsApp Number": ""})))
    A(_mut("Blank Product ID", "FAIL",
           "No product code and no default product selected -> row cannot form a line",
           lambda r, i: r.update({"Product ID": "", "Product Name": ""})))
    A(_mut("Unknown Product code", "FAIL", 'Unknown product code "PDT-99999"',
           lambda r, i: r.update({"Product ID": "PDT-99999",
                                  "Product Name": "Does Not Exist"})))
    A(_mut("Non-numeric Product code", "FAIL",
           "normalizeNumericCode() finds no digits -> unknown product code",
           lambda r, i: r.update({"Product ID": "PDT-ABC"})))
    A(_mut("Unknown Status label", "FAIL", 'Unknown status "Awaiting Warehouse"',
           lambda r, i: r.update({"Status": "Awaiting Warehouse"})))
    A(_mut("Typo Status", "FAIL", 'Unknown status "Delivred" (misspelt)',
           lambda r, i: r.update({"Status": "Delivred"})))
    A(_mut("Numeric Status", "FAIL", 'Unknown status "3" - CRM exported a status ID',
           lambda r, i: r.update({"Status": "3"})))
    A(_mut("Unknown Media Buyer code", "FAIL", 'Unknown media buyer code "USR-99999"',
           lambda r, i: r.update({"Media Buyer ID": "USR-99999",
                                  "Media-Buyer": "Ghost Buyer"})))
    A(_mut("MB name disagrees with MB ID", "IMPORT",
           "Name column says one person, ID says another. The importer IGNORES the "
           "name and attributes by ID. CHECK the order shows the USR-6 buyer, NOT "
           "the name printed on the sheet.",
           lambda r, i: r.update({"Media Buyer ID": "USR-6",
                                  "Media-Buyer": MB_BY_CODE["USR-7"]})))
    A(_mut("CS name disagrees with CS ID", "IMPORT",
           "Same check on the CS side: attribution must follow CS ID, not the CS name.",
           lambda r, i: r.update({"CS ID": "USR-13",
                                  "CS": CS_BY_CODE["USR-12"]})))
    A(_mut("MB name present, MB ID blank", "IMPORT",
           "A CRM export with names but no codes. The name alone attributes NOTHING, "
           "so the order falls back to System on the company's first branch.",
           lambda r, i: r.update({"Media Buyer ID": "", "Media-Buyer": "Exre Aluya"})))
    A(_mut("Non-numeric MB code", "FAIL", "No digits in code -> unknown media buyer",
           lambda r, i: r.update({"Media Buyer ID": "MB-UNKNOWN"})))
    A(_mut("Unknown CS code", "FAIL", 'Unknown CS code "USR-88888"',
           lambda r, i: r.update({"CS ID": "USR-88888", "CS": "Ghost Closer"})))
    A(_mut("Unknown Currency", "FAIL", 'Unknown or inactive currency "XYZ"',
           lambda r, i: r.update({"Currency": "XYZ"})))
    A(_mut("Unknown Country as currency", "FAIL",
           'Unknown or inactive currency "Atlantis"',
           lambda r, i: r.update({"Currency": "Atlantis"})))
    A(_mut("Inactive currency (USD)", "FAIL",
           "Fails unless USD is an ACTIVE currency in this company",
           lambda r, i: r.update({"Currency": "USD"})))
    A(_mut("Invalid Email format", "FAIL",
           "importOrderSchema requires a valid email when present",
           lambda r, i: r.update({"Email": "not-an-email"})))
    A(_mut("Email missing domain", "FAIL", "z.string().email() rejects 'joe@'",
           lambda r, i: r.update({"Email": "joe@"})))
    A(_mut("Negative Cost", "FAIL", "totalAmount must be >= 0",
           lambda r, i: r.update({"Cost": -50000})))
    A(_mut("Cost with 3 decimals", "FAIL",
           "totalAmount must be a multiple of 0.01 (money precision)",
           lambda r, i: r.update({"Cost": 15000.555})))
    A(_mut("Text in Cost", "FAIL",
           'Number("N/A") = NaN -> totalAmount fails number validation',
           lambda r, i: r.update({"Cost": "N/A"})))
    A(_mut("Cost 'FREE'", "FAIL", 'Number("FREE") = NaN',
           lambda r, i: r.update({"Cost": "FREE"})))
    A(_mut("Name over 100 chars in State", "FAIL",
           "deliveryState max 100 chars",
           lambda r, i: r.update({"State": "Lagos " * 30})))

    # ── CURRENCY-FORMAT COST: these look fine to a human but Number() rejects.
    A(_mut("Cost with Naira symbol", "FAIL",
           'Template says symbols tolerated, but Number("N100,000") = NaN. '
           "HIGH-VALUE BUG CHECK: paste-from-CRM exports look like this.",
           lambda r, i: r.update({"Cost": "N100,000"})))
    A(_mut("Cost with thousands comma", "FAIL",
           'Number("100,000") = NaN. Very common in CRM exports.',
           lambda r, i: r.update({"Cost": "100,000"})))
    A(_mut("Cost with trailing space+symbol", "FAIL",
           'Number("45,000 NGN") = NaN',
           lambda r, i: r.update({"Cost": "45,000 NGN"})))

    # ── SOFT / COERCED: should import, but verify the stored value ───────────
    A(_mut("Zero Cost", "IMPORT",
           "0 is allowed (min 0). Order total AND line item both 0.",
           lambda r, i: r.update({"Cost": 0})))
    A(_mut("Blank Cost", "IMPORT",
           "No Cost and no unit-price column -> line falls back to 0, so the order "
           "and its invoice both read 0. CHECK: is a money-less row acceptable?",
           lambda r, i: r.update({"Cost": ""})))
    A(_mut("Blank Quantity", "IMPORT", "Quantity defaults to 1",
           lambda r, i: r.update({"Quantity": ""})))
    A(_mut("Quantity 0", "IMPORT",
           "Number('0') is FALSY, so mapRow's `|| 1` turns it into 1. The order "
           "imports with qty 1, NOT 0. CHECK this is what you want: a 0 on the "
           "sheet silently becomes 1.",
           lambda r, i: r.update({"Quantity": 0})))
    A(_mut("Quantity decimal 2.5", "FAIL",
           "orderItemSchema requires quantity to be an INTEGER. 2.5 passes mapRow's "
           "coercion untouched and is rejected by Zod.",
           lambda r, i: r.update({"Quantity": 2.5})))
    A(_mut("Quantity negative", "FAIL",
           "orderItemSchema requires quantity >= 1. Number('-2') is truthy so it "
           "survives mapRow's `|| 1` fallback and is rejected by Zod.",
           lambda r, i: r.update({"Quantity": -2})))
    A(_mut("Quantity text", "IMPORT",
           'Number("two") = NaN which is falsy -> defaults to 1',
           lambda r, i: r.update({"Quantity": "two"})))
    A(_mut("Quantity huge", "IMPORT", "999999 units - check numeric overflow / stock",
           lambda r, i: r.update({"Quantity": 999999})))
    A(_mut("Cost huge", "IMPORT",
           "9,999,999,999.99 - check numeric(precision) does not overflow",
           lambda r, i: r.update({"Cost": 9999999999.99})))
    A(_mut("Blank Status", "FAIL",
           "Blank no longer inherits the job default (CEO 2026-09-01): the status must "
           "match a status in the app, so a blank cell fails and is fixed on the job page.",
           lambda r, i: r.update({"Status": ""})))
    A(_mut("Whitespace-only Status", "FAIL",
           "Trims to empty -> same as blank -> required.",
           lambda r, i: r.update({"Status": "   "})))
    A(_mut("Status lowercase", "IMPORT",
           "parseRowStatus lowercases -> 'delivered' maps to DELIVERED",
           lambda r, i: r.update({"Status": "delivered"})))
    A(_mut("Status UPPERCASE", "IMPORT", "'PENDING' maps to CS_ASSIGNED",
           lambda r, i: r.update({"Status": "PENDING"})))
    A(_mut("Status padded whitespace", "IMPORT", "Trimmed then matched",
           lambda r, i: r.update({"Status": "   Confirmed   "})))
    A(_mut("Status 'Delivered & Cash Remitted'", "IMPORT",
           "Accepted ampersand variant -> REMITTED",
           lambda r, i: r.update({"Status": "Delivered & Cash Remitted"})))
    A(_mut("Status 'No_Response'", "IMPORT", "Separator normalised -> CS_ENGAGED",
           lambda r, i: r.update({"Status": "No_Response"})))
    A(_mut("Status 'Canceled' (US spelling)", "IMPORT",
           "One-l variant is an accepted label -> DELETED",
           lambda r, i: r.update({"Status": "Canceled"})))
    A(_mut("Status 'Order Delivered'", "FAIL",
           "Not an exact accepted label. Strict matching (CEO 2026-09-01) rejects it "
           "rather than guessing DELIVERED.",
           lambda r, i: r.update({"Status": "Order Delivered"})))
    A(_mut("Status 'Not Delivered'", "FAIL",
           "CRITICAL: the old substring matcher imported this as DELIVERED, the "
           "OPPOSITE meaning, inflating delivered counts that feed COGS and remittance. "
           "Must now FAIL.",
           lambda r, i: r.update({"Status": "Not Delivered"})))
    A(_mut("Status 'Delivery Failed'", "FAIL",
           "Same class of bug as 'Not Delivered': must never import as DELIVERED.",
           lambda r, i: r.update({"Status": "Delivery Failed"})))
    A(_mut("Status 'Pending Cancellation'", "FAIL",
           "Old matcher hit 'pending' first and imported as CS_ASSIGNED. Now rejected.",
           lambda r, i: r.update({"Status": "Pending Cancellation"})))
    A(_mut("Status 'Cancelled' -> DELETED", "IMPORT",
           "CANCELLED is legacy-only (CLAUDE.md), so a cancelled source row imports "
           "as DELETED. CHECK it is excluded from every metric.",
           lambda r, i: r.update({"Status": "Cancelled"})))
    A(_mut("Blank MB + blank CS", "IMPORT",
           "Both blank -> branch falls back to company first branch, order unattributed (System)",
           lambda r, i: r.update({"Media Buyer ID": "", "CS ID": "",
                                  "Media-Buyer": "", "CS": ""})))
    A(_mut("Blank MB only", "IMPORT", "Branch derived from the CS closer instead",
           lambda r, i: r.update({"Media Buyer ID": "", "Media-Buyer": ""})))
    A(_mut("Blank CS only", "IMPORT", "Branch derived from the media buyer",
           lambda r, i: r.update({"CS ID": "", "CS": ""})))
    A(_mut("Bare numeric MB code", "IMPORT",
           'normalizeNumericCode strips prefixes: "5" resolves same as "USR-5"',
           lambda r, i: r.update({"Media Buyer ID": "5",
                                  "Media-Buyer": MB_BY_CODE["USR-5"]})))
    A(_mut("Zero-padded MB code", "IMPORT", '"USR-005" -> leading zeros stripped -> 5',
           lambda r, i: r.update({"Media Buyer ID": "USR-005",
                                  "Media-Buyer": MB_BY_CODE["USR-5"]})))
    A(_mut("MB code no separator", "IMPORT", '"USR5" resolves to 5',
           lambda r, i: r.update({"Media Buyer ID": "USR5",
                                  "Media-Buyer": MB_BY_CODE["USR-5"]})))
    A(_mut("Product code bare number", "IMPORT", '"1" resolves same as "PDT-1"',
           lambda r, i: r.update({"Product ID": "1"})))
    A(_mut("Product code lowercase", "IMPORT", '"pdt-2" -> digits extracted -> 2',
           lambda r, i: r.update({"Product ID": "pdt-2"})))
    A(_mut("Currency lowercase", "IMPORT", '"ngn" upper-cased before lookup',
           lambda r, i: r.update({"Currency": "ngn"})))
    A(_mut("Currency as country name", "IMPORT",
           '"Nigeria" resolves via the country-name key',
           lambda r, i: r.update({"Currency": "Nigeria"})))
    A(_mut("Blank Currency", "IMPORT",
           "Falls back to MB country -> CS country -> selected country -> base",
           lambda r, i: r.update({"Currency": ""})))

    # ── DATE handling ────────────────────────────────────────────────────────
    A(_mut("ISO date", "IMPORT", "2026-03-15T10:30:00Z accepted",
           lambda r, i: r.update({"Date": "2026-03-15T10:30:00Z"})))
    A(_mut("US date with time", "IMPORT", '"5/2/2026 9:05:30" per the template',
           lambda r, i: r.update({"Date": "5/2/2026 9:05:30"})))
    A(_mut("Ambiguous DD/MM date", "IMPORT",
           "13/4/2026 - only valid as DD/MM. CHECK: is it read as 13 Apr or rejected/shifted?",
           lambda r, i: r.update({"Date": "13/4/2026"})))
    A(_mut("Dotted date", "IMPORT", "15.03.2026 - European separator",
           lambda r, i: r.update({"Date": "15.03.2026"})))
    A(_mut("Blank Date", "IMPORT", "Defaults to import time (today)",
           lambda r, i: r.update({"Date": ""})))
    A(_mut("Garbage Date", "IMPORT",
           "'not a date' - invalid Date. CHECK it falls back to today, not 1970 or crash.",
           lambda r, i: r.update({"Date": "not a date"})))
    A(_mut("Far-future Date", "IMPORT", "Year 2099 - check it is not silently clamped",
           lambda r, i: r.update({"Date": "1/1/2099"})))
    A(_mut("Far-past Date", "IMPORT", "Year 1970 - epoch boundary",
           lambda r, i: r.update({"Date": "1/1/1970"})))
    A(_mut("Excel serial date number", "IMPORT",
           "46000 - a raw Excel serial that lost its date format on export",
           lambda r, i: r.update({"Date": 46000})))

    # ── PHONE formats (Pillar 2: none of these may appear raw in responses) ──
    A(_mut("Phone with spaces", "IMPORT", "Template says spaces tolerated",
           lambda r, i: r.update({"Phone Number": "080 6888 0766"})))
    A(_mut("Phone with dashes", "IMPORT", "Dashes tolerated",
           lambda r, i: r.update({"Phone Number": "080-6888-0766"})))
    A(_mut("Phone +234 intl", "IMPORT", "International format",
           lambda r, i: r.update({"Phone Number": "+2348068880766"})))
    A(_mut("Phone with parentheses", "IMPORT", "(080) 6888 0766",
           lambda r, i: r.update({"Phone Number": "(080) 6888 0766"})))
    A(_mut("Phone stored as number", "IMPORT",
           "8068880766 as a NUMBER - Excel dropped the leading zero",
           lambda r, i: r.update({"Phone Number": 8068880766})))
    A(_mut("Phone in scientific notation", "IMPORT",
           "8.06888E+09 - classic Excel long-number corruption. "
           "CHECK the stored phone is not garbage.",
           lambda r, i: r.update({"Phone Number": "8.06888E+09"})))
    A(_mut("Phone too long (over 50)", "FAIL",
           "customerPhone max 50 chars",
           lambda r, i: r.update({"Phone Number": "0" * 60})))
    A(_mut("Phone with letters", "IMPORT",
           '"CALL SHOP" - no format validation, so it imports. '
           "CHECK: do you want this rejected?",
           lambda r, i: r.update({"Phone Number": "CALL SHOP"})))

    # ── TEXT / ENCODING ──────────────────────────────────────────────────────
    A(_mut("Unicode name", "IMPORT", "Adébáyò Ògúnlànà - accents must survive round-trip",
           lambda r, i: r.update({"Name": "Adébáyò Ògúnlànà"})))
    A(_mut("Arabic name", "IMPORT", "RTL script must not corrupt",
           lambda r, i: r.update({"Name": "محمد الأمين"})))
    A(_mut("Emoji in name", "IMPORT", "Check emoji storage/display",
           lambda r, i: r.update({"Name": "Blessing 🎉 Okon"})))
    A(_mut("Very long name", "IMPORT", "300-char name - check DB column limit",
           lambda r, i: r.update({"Name": "Chukwuemeka " * 25})))
    A(_mut("Very long address", "IMPORT", "1000-char address",
           lambda r, i: r.update({"Address": "Plot 15 Admiralty Way Lekki Phase 1 " * 28})))
    A(_mut("SQL-injection-looking name", "IMPORT",
           "Robert'); DROP TABLE orders;-- must be stored literally, never executed",
           lambda r, i: r.update({"Name": "Robert'); DROP TABLE orders;--"})))
    A(_mut("HTML/XSS in name", "IMPORT",
           "<script>alert(1)</script> must render escaped, never execute in the UI",
           lambda r, i: r.update({"Name": "<script>alert(1)</script>"})))
    A(_mut("Formula injection", "IMPORT",
           '=1+1 in a cell. CHECK: exported CSV must not re-execute it in Excel.',
           lambda r, i: r.update({"Name": "=1+1", "Address": "=cmd|'/c calc'!A1"})))
    A(_mut("Leading/trailing spaces everywhere", "IMPORT",
           "readCell trims - verify no double spaces stored",
           lambda r, i: r.update({"Name": "  Ngozi Okafor  ",
                                  "State": "  Lagos  ",
                                  "Product ID": " PDT-1 ",
                                  "Media Buyer ID": " USR-5 ",
                                  "Media-Buyer": MB_BY_CODE["USR-5"]})))
    A(_mut("Tabs and newlines in address", "IMPORT",
           "Multi-line address from a textarea export",
           lambda r, i: r.update({"Address": "12 Awolowo Road\nApt 4B\tIkoyi"})))
    A(_mut("Comma-heavy address (CSV bomb)", "IMPORT",
           "If saved as CSV this must stay one field, not split into columns",
           lambda r, i: r.update({"Address": "12, Awolowo Road, Ikoyi, Lagos, Nigeria"})))
    A(_mut("Quotes in address", "IMPORT",
           'Double quotes must survive CSV round-trip',
           lambda r, i: r.update({"Address": 'The "Blue Gate" House, Lekki'})))
    A(_mut("All optional fields blank", "IMPORT",
           "Only the required fields present - the minimum viable row",
           lambda r, i: r.update({"WhatsApp Number": "", "Email": "", "Address": "",
                                  "State": "", "Gender": "", "Delivery Time": "",
                                  "More details": "", "Delivery agent": "",
                                  "Comment 1": "", "Comment 2": "", "Comment 3": "",
                                  "Product Name": ""})))
    A(_mut("All comments filled", "IMPORT",
           "Comments 1-3 combined into custom fields",
           lambda r, i: r.update({"Comment 1": "Called twice, no answer",
                                  "Comment 2": "Rescheduled to Friday",
                                  "Comment 3": "Customer confirmed address"})))
    A(_mut("Product Name mismatches Product ID", "IMPORT",
           "Product Name is reference-only and IGNORED. "
           "CHECK the order shows the PDT-1 product, not 'Totally Different Product'.",
           lambda r, i: r.update({"Product ID": "PDT-1",
                                  "Product Name": "Totally Different Product"})))
    A(_mut("Gender junk value", "IMPORT", "'Yes' in the gender column - free text, imports",
           lambda r, i: r.update({"Gender": "Yes"})))


    # ── MONEY / LINE-ITEM COHERENCE (unitPrice falls back to Cost) ───────────
    A(_mut("Cost with decimals (.50)", "IMPORT",
           "2 decimal places is valid money. Line item and invoice must both show "
           "the same value as the order total.",
           lambda r, i: r.update({"Cost": 47500.50})))
    A(_mut("Cost 0.01 (minimum money)", "IMPORT",
           "Smallest valid non-zero amount; checks numeric rounding.",
           lambda r, i: r.update({"Cost": 0.01})))
    A(_mut("Cost 0.005 (sub-kobo)", "FAIL",
           "Below kobo precision: not a multiple of 0.01.",
           lambda r, i: r.update({"Cost": 0.005})))
    A(_mut("High qty + Cost (line total check)", "IMPORT",
           "CRITICAL: unitPrice IS the LINE TOTAL, never multiplied by qty. "
           "Cost 175000 with qty 5 must show ONE line of 175000, NOT 875000. "
           "This is the case that was importing as 0 before the fix.",
           lambda r, i: r.update({"Quantity": 5, "Cost": 175000})))
    A(_mut("Cost in scientific notation", "IMPORT",
           "1.75E+05 -> Number() reads 175000. CHECK the stored amount is right.",
           lambda r, i: r.update({"Cost": "1.75E+05"})))
    A(_mut("Cost with trailing/leading space", "IMPORT",
           "' 45000 ' trims to a clean number.",
           lambda r, i: r.update({"Cost": " 45000 "})))
    A(_mut("Cost as accounting negative", "FAIL",
           "'(45000)' is accounting notation for negative. Number() = NaN.",
           lambda r, i: r.update({"Cost": "(45000)"})))

    # ── DUPLICATE / IDEMPOTENCY within the same file ─────────────────────────
    A(_mut("Order ID with leading apostrophe", "IMPORT",
           "\"'CRM-9001\" - Excel text-marker apostrophe. CHECK it is not stored "
           "as part of the external id (would break re-import idempotency).",
           lambda r, i: r.update({"Order ID": "'CRM-90" + str(i)})))
    A(_mut("Order ID with inner spaces", "IMPORT",
           "'CRM 1234' - trimmed at the ends only, inner space is kept as-is.",
           lambda r, i: r.update({"Order ID": f"CRM {i}"})))
    A(_mut("Very long Order ID", "IMPORT",
           "300-char external id; checks the column and the unique index hold.",
           lambda r, i: r.update({"Order ID": f"CRM-{i}-" + "X" * 300})))

    # ── PRODUCT edge cases ───────────────────────────────────────────────────
    A(_mut("Product code with spaces inside", "IMPORT",
           "'PDT - 1' -> digits extracted -> resolves to product 1.",
           lambda r, i: r.update({"Product ID": "PDT - 1"})))
    A(_mut("Product code 0", "FAIL",
           "'PDT-0' normalises to empty (leading zeros stripped) -> unknown product.",
           lambda r, i: r.update({"Product ID": "PDT-0"})))
    A(_mut("Product code negative", "IMPORT",
           "QUIRK: '-1' resolves to PDT-1 because normalizeNumericCode strips "
           "everything non-numeric, sign included. It imports silently. Decide "
           "whether a negative code should be rejected instead.",
           lambda r, i: r.update({"Product ID": "-1"})))
    A(_mut("Product code with decimal point", "DEPENDS",
           "DANGEROUS QUIRK: 'PDT-1.0' strips the dot -> code '10'. If product 10 "
           "EXISTS the row imports the WRONG product silently; if not, it fails as "
           "an unknown code. Excel turning 1 into 1.0 is exactly how this happens. "
           "Either outcome is a bug worth deciding on.",
           lambda r, i: r.update({"Product ID": "PDT-1.0"})))
    A(_mut("MB code with decimal point", "DEPENDS",
           "Same quirk on the user side: 'USR-5.0' -> code '50'. Imports under the "
           "WRONG media buyer if user 50 exists, else fails. Attribution is never "
           "USR-5 either way.",
           lambda r, i: r.update({"Media Buyer ID": "USR-5.0"})))
    A(_mut("Product Name filled, Product ID blank", "FAIL",
           "A CRM export with names but no codes. Product Name is IGNORED, so the "
           "row has no product and fails. Very likely in a real migration.",
           lambda r, i: r.update({"Product ID": "", "Product Name": "Slim Tea Pack"})))

    # ── STATUS x LIFECYCLE coherence ─────────────────────────────────────────
    A(_mut("REMITTED with no CS assigned", "IMPORT",
           "Historical completed order with a blank CS. CHECK it still lands in "
           "Cash Remittances and is not stranded without an owner.",
           lambda r, i: r.update({"Status": "Delivered and Cash Remitted",
                                  "CS ID": "", "CS": ""})))
    A(_mut("DELIVERED with no MB and no CS", "IMPORT",
           "Fully unattributed delivered order: branch falls back to the company's "
           "first branch and the order imports as System. CHECK which branch it "
           "lands in, and that delivered metrics still count it.",
           lambda r, i: r.update({"Status": "Delivered", "Media Buyer ID": "",
                                  "Media-Buyer": "", "CS ID": "", "CS": ""})))
    A(_mut("Cancelled -> DELETED exclusion", "IMPORT",
           "Imports as DELETED (CANCELLED is legacy-only). CHECK it is excluded "
           "from EVERY metric, not merely hidden from the default list.",
           lambda r, i: r.update({"Status": "Cancelled"})))
    A(_mut("Returned order with a cost", "IMPORT",
           "RETURNED carrying money. CHECK stock and COGS treatment on a returned "
           "historical order (returns are meant to be stock-neutral).",
           lambda r, i: r.update({"Status": "Returned", "Cost": 62000})))

    # ── BRANCH derivation (branch is NEVER on the sheet) ─────────────────────
    A(_mut("MB and CS in different branches", "IMPORT",
           "CRITICAL: branch comes from the MEDIA BUYER first, and import sets "
           "branch_id AND servicing_branch_id to the SAME value. So the CS closer's "
           "branch is ignored. CHECK which branch this lands in, and whether the "
           "CS's own branch pages can still see it.",
           lambda r, i: r.update({"Media Buyer ID": "USR-5",
                                  "Media-Buyer": MB_BY_CODE["USR-5"],
                                  "CS ID": "USR-14",
                                  "CS": CS_BY_CODE["USR-14"]})))

    # ── CURRENCY / COUNTRY SCOPE ─────────────────────────────────────────────
    A(_mut("Currency with whitespace", "IMPORT",
           "' NGN ' trims and upper-cases before lookup.",
           lambda r, i: r.update({"Currency": " ngn "})))
    A(_mut("Currency code too long", "FAIL",
           "'NIGERIAN' is not an active currency code or country name.",
           lambda r, i: r.update({"Currency": "NIGERIAN"})))

    # ── DATE edge cases not yet covered ──────────────────────────────────────
    A(_mut("Date only (no time)", "IMPORT",
           "'2026-03-15' with no time component.",
           lambda r, i: r.update({"Date": "2026-03-15"})))
    A(_mut("Date with timezone offset", "IMPORT",
           "'2026-03-15T10:30:00+01:00' (WAT). CHECK the stored date does not "
           "shift a day when rendered.",
           lambda r, i: r.update({"Date": "2026-03-15T10:30:00+01:00"})))
    A(_mut("Feb 30 (impossible date)", "IMPORT",
           "'2/30/2026' does not exist. CHECK it falls back to today rather than "
           "rolling over to 2 March.",
           lambda r, i: r.update({"Date": "2/30/2026"})))
    A(_mut("Two-digit year", "IMPORT",
           "'5/2/26' - CHECK it reads as 2026 and not 1926.",
           lambda r, i: r.update({"Date": "5/2/26"})))
    A(_mut("Date as pure number string", "IMPORT",
           "'20260315' - compact form; CHECK how it parses.",
           lambda r, i: r.update({"Date": "20260315"})))

    # ── EMAIL / STATE / GENDER field limits ──────────────────────────────────
    A(_mut("Email with spaces", "FAIL",
           "'joe smith@example.com' is not a valid email.",
           lambda r, i: r.update({"Email": "joe smith@example.com"})))
    A(_mut("Email over 255 chars", "FAIL",
           "customerEmail max 255.",
           lambda r, i: r.update({"Email": ("e" * 250) + "@example.com"})))
    A(_mut("Email uppercase", "IMPORT",
           "'JOE@EXAMPLE.COM' is valid; CHECK storage casing is consistent.",
           lambda r, i: r.update({"Email": "JOE@EXAMPLE.COM"})))
    A(_mut("State exactly 100 chars", "IMPORT",
           "Boundary: max is 100, so exactly 100 must PASS (101 fails).",
           lambda r, i: r.update({"State": "L" * 100})))
    A(_mut("Phone exactly 50 chars", "IMPORT",
           "Boundary: max is 50, so exactly 50 must PASS (51 fails).",
           lambda r, i: r.update({"Phone Number": "0" * 50})))
    A(_mut("Name exactly 2 chars", "IMPORT",
           "Boundary: min is 2, so 2 must PASS (1 fails).",
           lambda r, i: r.update({"Name": "Jo"})))

    return M


def build_rows(total, seed=42):
    rng = random.Random(seed)
    base = datetime(2026, 1, 1)
    mutators = build_mutators()
    rows, cases = [], []

    # Every mutator appears at least 3 times, spread across the file, so the
    # CEO sees each failure mode more than once and at different offsets
    # (also exercises chunk boundaries at 500 rows).
    REPEATS = 3
    planned = []
    for rep in range(REPEATS):
        for mi, m in enumerate(mutators):
            planned.append(mi)
    rng.shuffle(planned)

    if len(planned) > total:
        planned = planned[:total]

    # Place the edge cases at pseudo-random positions across the whole file.
    # Rows reserved for the multi-row structural cases below are excluded so a
    # mutator can never land on them and contradict their stated expectation.
    RESERVED = {5, 6, 10, 11, 15, 20, 21, 25, 26, 27, 498, 499, 500}
    available = [i for i in range(total) if i not in RESERVED]
    if len(planned) > len(available):
        planned = planned[: len(available)]
    positions = sorted(rng.sample(available, len(planned)))
    at = dict(zip(positions, planned))

    for i in range(total):
        n = 1001 + i
        row = make_good_row(rng, n, base)
        if i in at:
            m = mutators[at[i]]
            m["fn"](row, n)
            cases.append({
                "row": i + 2,  # +2: 1-based, plus header row
                "orderId": row["Order ID"],
                "case": m["name"],
                "expect": m["expect"],
                "reason": m["reason"],
            })
        rows.append(row)

    # ── Structural cases that need to span MULTIPLE rows ─────────────────────
    # Duplicate Order IDs: idempotency key collision inside one file.
    if total >= 40:
        a, b = 5, 6
        rows[b] = dict(rows[a])
        rows[b]["Name"] = rows[b]["Name"] + " (second copy)"
        rows[b]["Cost"] = 77777
        cases.append({"row": b + 2, "orderId": rows[b]["Order ID"],
                      "case": "DUPLICATE Order ID (same file)",
                      "expect": "IMPORT",
                      "reason": f"Same Order ID as row {a + 2}. Upsert means the LAST row wins - "
                                "expect ONE order with cost 77777, not two orders."})

        # Exact duplicate row (every column identical).
        c, d = 10, 11
        rows[d] = dict(rows[c])
        cases.append({"row": d + 2, "orderId": rows[d]["Order ID"],
                      "case": "EXACT duplicate row",
                      "expect": "IMPORT",
                      "reason": f"Byte-identical to row {c + 2}. Must produce ONE order, no duplicate."})

        # Fully empty row in the middle of the file.
        e = 15
        rows[e] = {h: "" for h in HEADERS}
        cases.append({"row": e + 2, "orderId": "(blank)",
                      "case": "COMPLETELY EMPTY row",
                      "expect": "FAIL",
                      "reason": "Blank line in the middle of the sheet. Must fail cleanly "
                                "(missing external id), not crash or halt the job."})

        # Case-differing duplicate ID.
        f, g = 20, 21
        rows[g]["Order ID"] = rows[f]["Order ID"].lower()
        cases.append({"row": g + 2, "orderId": rows[g]["Order ID"],
                      "case": "Order ID differing only by CASE",
                      "expect": "IMPORT",
                      "reason": f"'crm-{rows[f]['Order ID'][4:]}' vs row {f + 2}'s uppercase. "
                                "CHECK: two orders, or one? Reveals whether the key is case-sensitive."})

        # Same customer phone across many rows (repeat customer, not a dup).
        rep_phone = "08033344455"
        for idx in (25, 26, 27):
            rows[idx]["Phone Number"] = rep_phone
            rows[idx]["WhatsApp Number"] = rep_phone
        cases.append({"row": 27, "orderId": rows[25]["Order ID"],
                      "case": "Same phone, 3 different orders",
                      "expect": "IMPORT",
                      "reason": "Rows 27-29 share one phone with distinct Order IDs. "
                                "Import skips dedup, so expect 3 separate orders."})

        # Chunk-boundary rows: CHUNK_SIZE is 500.
        for idx in (498, 499, 500):
            if idx < total:
                rows[idx]["More details"] = "CHUNK BOUNDARY MARKER"
                cases.append({"row": idx + 2, "orderId": rows[idx]["Order ID"],
                              "case": "Chunk boundary (CHUNK_SIZE=500)",
                              "expect": "IMPORT",
                              "reason": "Sits on the 500-row chunk edge. Verify no row is "
                                        "skipped or double-counted at the boundary."})

    cases.sort(key=lambda c: c["row"])
    return rows, cases


def reference_rows():
    return [
        ("Order ID", "Required. A unique ID from your source (the idempotency key). "
                     "Re-importing the same ID OVERWRITES that order instead of creating a duplicate."),
        ("Date", 'Order date. Accepts "4/29/2026", "5/2/2026 9:05:30", or ISO format. '
                 "Optional: defaults to today."),
        ("Name", "Customer name. Required, min 2 characters."),
        ("Phone Number", "Customer phone. Required. Accepts any format (spaces, dashes tolerated). Max 50 chars."),
        ("WhatsApp Number", "Optional. Stored as reference in custom fields."),
        ("Email", "Optional. Must be a valid email if provided."),
        ("Address", "Customer / delivery address. Optional."),
        ("State", "Delivery state (e.g. Lagos, Abuja, Rivers). Optional. Max 100 chars."),
        ("Product ID", "Required. Product code (e.g. PDT-1). An unknown code fails the row. "
                       "Only the DIGITS are read, so PDT-1, 1 and 001 are the same product. "
                       "Beware: PDT-1.0 reads as 10 (a different product)."),
        ("Product Name", "Optional reference only. The import IGNORES this column and matches on Product ID."),
        ("Quantity", "Whole number, 1 or more. Blank, 0 or text all become 1. "
                     "A decimal (2.5) or a negative fails the row."),
        ("Cost", "Order total. Must be a plain number: 100000, at most 2 decimals. "
                 "WARNING: commas and currency symbols (100,000 or N100,000) FAIL the row. "
                 "When the file has no separate unit-price column, Cost also becomes the "
                 "line-item price, so the order total and its invoice match."),
        ("Currency", "Optional. Currency code or country name (e.g. NGN, GHS, Ghana). "
                     "If blank, uses the media buyer / closer country, then the base currency. "
                     "An unknown or inactive currency fails the row."),
        ("Gender", "Optional (e.g. Male, Female)."),
        ("Delivery Time", "Optional free text (e.g. Tomorrow, 3 Days, Today)."),
        ("More details", "Optional notes about delivery."),
        ("Status", "Required on every row. Must EXACTLY match a status below. "
                   "A blank cell or an unrecognised label fails the row."),
        ("Media-Buyer", "Optional. Stored as reference (name). The Media Buyer ID column drives attribution."),
        ("Media Buyer ID", "Optional. User code (e.g. USR-5). An unknown code fails the row. "
                           "Only the DIGITS are read (USR-5, 5 and 005 all match). "
                           "THE ORDER'S BRANCH COMES FROM THIS USER: the media buyer's branch "
                           "first, then the CS closer's, then the company's first branch "
                           "(in which case the order imports unattributed, as System). "
                           "There is no branch column on this sheet."),
        ("CS", "Optional. Stored as reference (name). The CS ID column drives assignment."),
        ("CS ID", "Optional. User code (e.g. USR-12). An unknown code fails the row."),
        ("Delivery agent", "Optional. Stored as reference in custom fields."),
        ("Comment 1-3", "Optional. Combined and stored in custom fields."),
        ("", ""),
        ("Valid statuses", "Must match EXACTLY (case and spacing ignored). "
                           "Anything else fails the row, and a blank Status fails too."),
        ("Unprocessed", "Imported as UNPROCESSED"),
        ("Pending", "Imported as CS_ASSIGNED"),
        ("No Response", "Imported as CS_ENGAGED"),
        ("Rescheduled", "Imported as CS_ENGAGED"),
        ("Confirmed", "Imported as CONFIRMED"),
        ("Delivered", "Imported as DELIVERED"),
        ("Delivered and Cash Remitted", "Imported as REMITTED"),
        ("Remitted", "Imported as REMITTED"),
        ("Returned", "Imported as RETURNED"),
        ("Cancelled", "Imported as DELETED (CANCELLED is legacy-only)"),
        ("", ""),
        ("NOT accepted", "These FAIL: 'Not Delivered', 'Delivery Failed', 'Order Delivered', "
                         "'Pending Cancellation', 'Delivred', or any blank cell."),
    ]


def write_workbook(path, rows, cases):
    wb = Workbook()

    hdr_fill = PatternFill("solid", fgColor="1F2937")
    hdr_font = Font(color="FFFFFF", bold=True)

    # ── Sheet 1: Orders ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Orders"
    ws.append(HEADERS)
    for c in ws[1]:
        c.fill, c.font = hdr_fill, hdr_font
    for r in rows:
        ws.append([r.get(h, "") for h in HEADERS])
    widths = [14, 20, 26, 18, 18, 30, 46, 14, 12, 26, 10, 16, 12, 10, 15, 30,
              28, 16, 16, 14, 12, 18, 34, 30, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Force text-ish columns to plain text so Excel does not "helpfully" reformat
    # phones into scientific notation on open.
    for col in ("D", "E"):
        for cell in ws[col][1:]:
            if isinstance(cell.value, str):
                cell.number_format = "@"

    # ── Sheet 2: Reference ───────────────────────────────────────────────────
    ref = wb.create_sheet("Reference")
    ref.append(["Column", "Rule"])
    for c in ref[1]:
        c.fill, c.font = hdr_fill, hdr_font
    for a, b in reference_rows():
        ref.append([a, b])
    ref.column_dimensions["A"].width = 30
    ref.column_dimensions["B"].width = 96
    for row in ref.iter_rows(min_row=2):
        row[1].alignment = Alignment(wrap_text=True, vertical="top")
    ref.freeze_panes = "A2"

    # ── Sheet 3: Test Cases ──────────────────────────────────────────────────
    tc = wb.create_sheet("Test Cases")
    tc.append(["Sheet Row", "Order ID", "Edge Case", "Expected", "Why / What to check"])
    for c in tc[1]:
        c.fill, c.font = hdr_fill, hdr_font
    fail_fill = PatternFill("solid", fgColor="FEE2E2")
    ok_fill = PatternFill("solid", fgColor="DCFCE7")
    for c in cases:
        tc.append([c["row"], c["orderId"], c["case"], c["expect"], c["reason"]])
        last = tc.max_row
        warn_fill = PatternFill("solid", fgColor="FEF3C7")
        tc.cell(last, 4).fill = (
            fail_fill if c["expect"] == "FAIL"
            else warn_fill if c["expect"] == "DEPENDS"
            else ok_fill
        )
        tc.cell(last, 4).font = Font(bold=True)
        tc.cell(last, 5).alignment = Alignment(wrap_text=True, vertical="top")
    for col, w in zip("ABCDE", (11, 14, 38, 11, 88)):
        tc.column_dimensions[col].width = w
    tc.freeze_panes = "A2"

    wb.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--rows", type=int, default=1000)
    ap.add_argument("--out", required=True)
    ap.add_argument("--seed", type=int, default=42)
    args = ap.parse_args()

    rows, cases = build_rows(args.rows, args.seed)
    write_workbook(args.out, rows, cases)

    fails = sum(1 for c in cases if c["expect"] == "FAIL")
    depends = sum(1 for c in cases if c["expect"] == "DEPENDS")
    imports = len(cases) - fails - depends
    print(f"Wrote {args.out}")
    print(f"  Orders sheet : {len(rows)} data rows")
    print(f"  Edge cases   : {len(cases)} tagged rows "
          f"({fails} FAIL, {imports} IMPORT, {depends} DEPENDS)")
    print(f"  Clean rows   : {len(rows) - len(cases)}")


if __name__ == "__main__":
    main()
